import sqlite3
from pathlib import Path

from openpyxl import Workbook

from app.services.hs_codes import (
    build_hs_generate_preview,
    parse_hs_code_upload_workbook,
    save_hs_upload_preview,
)


def test_parse_hs_code_upload_supports_aliases(tmp_path: Path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "GTS summary title"
    worksheet["B3"] = "GTS"
    worksheet["D3"] = "hscode"
    worksheet["B4"] = "GTS-HS-001"
    worksheet["D4"] = 87089910
    path = tmp_path / "hs-alias.xlsx"
    workbook.save(path)

    rows = parse_hs_code_upload_workbook(path)

    assert len(rows) == 1
    assert rows[0].row_number == 4
    assert rows[0].values["gts_no_normalized"] == "GTSHS001"
    assert rows[0].values["hs_code"] == "87089910"


def test_parse_hs_code_upload_ignores_rows_without_gts_or_hs_code(tmp_path: Path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "GTS"
    worksheet["B1"] = "HS Code"
    worksheet["A2"] = ""
    worksheet["B2"] = "87089910"
    worksheet["A3"] = "GTS-HS-002"
    worksheet["B3"] = ""
    worksheet["A4"] = "GTS-HS-003"
    worksheet["B4"] = "87089930"
    path = tmp_path / "hs-skip-empty.xlsx"
    workbook.save(path)

    rows = parse_hs_code_upload_workbook(path)

    assert len(rows) == 1
    assert rows[0].row_number == 4
    assert rows[0].values["gts_no_normalized"] == "GTSHS003"
    assert rows[0].values["hs_code"] == "87089930"


def test_build_hs_generate_preview_preserves_request_order():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    connection.execute(
        """
        CREATE TABLE products (
            id INTEGER PRIMARY KEY,
            gts_no TEXT,
            gts_no_normalized TEXT,
            oem TEXT,
            hs_code TEXT,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_by TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE quotation_items (
            id INTEGER PRIMARY KEY,
            product_id INTEGER NOT NULL,
            gts_no_normalized TEXT,
            updated_at TEXT NOT NULL
        )
        """
    )
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, oem, hs_code,
            created_by, created_at, updated_by, updated_at
        )
        VALUES
            (1, 'GTS-1', 'GTS1', 'OEM-1', '1111', 'A', '2026-01-01', 'A', '2026-01-01'),
            (2, 'GTS-2', 'GTS2', 'OEM-2', '2222', 'A', '2026-01-01', 'A', '2026-01-01')
        """
    )
    rows = [
        make_hs_row(5, "GTS-2"),
        make_hs_row(6, "GTS-1"),
    ]

    preview = build_hs_generate_preview(connection, rows)

    assert [row["values"]["gts_no"] for row in preview] == ["GTS-2", "GTS-1"]
    assert [row["product"]["hs_code"] for row in preview] == ["2222", "1111"]


def test_save_hs_upload_preview_does_not_update_direct_payload_without_hs_code():
    connection = build_hs_connection()
    preview_rows = [
        {
            "row_number": 2,
            "values": {
                "gts_no": "GTS-1",
                "gts_no_normalized": "GTS1",
                "hs_code": "",
            },
            "errors": [],
            "status": "matched",
            "product": {"id": 1},
        }
    ]

    result = save_hs_upload_preview(
        connection,
        preview_rows=preview_rows,
        operator_name="Nancy",
        file_name="hs.xlsx",
    )

    product = connection.execute("SELECT hs_code FROM products WHERE id = 1").fetchone()
    assert result == {"updated": 0, "failed": 1}
    assert product["hs_code"] == "1111"


def make_hs_row(row_number, gts_no):
    from app.services.hs_codes import ParsedHsCodeRow

    return ParsedHsCodeRow(
        row_number=row_number,
        values={"gts_no": gts_no, "gts_no_normalized": gts_no.replace("-", "")},
        warnings=[],
        errors=[],
    )


def build_hs_connection():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    connection.execute(
        """
        CREATE TABLE products (
            id INTEGER PRIMARY KEY,
            gts_no TEXT,
            gts_no_normalized TEXT,
            oem TEXT,
            hs_code TEXT,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_by TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE operation_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action_time TEXT NOT NULL,
            operator_name TEXT NOT NULL,
            action_type TEXT NOT NULL,
            file_name TEXT,
            row_count INTEGER,
            note TEXT
        )
        """
    )
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, oem, hs_code,
            created_by, created_at, updated_by, updated_at
        )
        VALUES (1, 'GTS-1', 'GTS1', 'OEM-1', '1111', 'A', '2026-01-01', 'A', '2026-01-01')
        """
    )
    return connection
