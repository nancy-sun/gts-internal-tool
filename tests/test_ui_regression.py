from __future__ import annotations

import sqlite3
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook


ACCESS_CODE = "test-access-code"


@pytest.fixture()
def ui_client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> TestClient:
    database_path = tmp_path / "ui-regression.sqlite3"
    upload_path = tmp_path / "uploads"
    upload_path.mkdir()

    monkeypatch.setenv("SHARED_ACCESS_CODE", ACCESS_CODE)
    monkeypatch.setenv("SESSION_SECRET_KEY", "test-session-secret-key")
    monkeypatch.setenv("PRODUCT_EDIT_PASSWORD", "55123511")
    monkeypatch.setenv("DATABASE_PATH", str(database_path))

    from app.config import get_settings
    from app.main import create_app
    from app.routes import generate, hs_codes, upload

    get_settings.cache_clear()
    monkeypatch.setattr(upload, "UPLOAD_DIR", upload_path)
    monkeypatch.setattr(generate, "UPLOAD_DIR", upload_path)
    monkeypatch.setattr(hs_codes, "UPLOAD_DIR", upload_path)

    client = TestClient(create_app())
    client.database_path = database_path
    client.upload_path = upload_path
    login_response = client.post("/login", data={"access_code": ACCESS_CODE})
    assert login_response.status_code == 200
    return client


def test_layout_css_keeps_table_pages_stable() -> None:
    css = Path("app/static/styles.css").read_text(encoding="utf-8")

    assert ".page:has(.preview-page-form)" in css
    assert ".preview-page-form .table-wrap" in css
    assert ".upload-preview-table" in css
    assert "min-width: 1180px" in css
    assert ".generation-preview-table" in css
    assert ".search-results-table" in css
    assert ".data-quality-table" in css
    assert ".table-section + .table-section" in css
    assert "input.is-invalid" in css
    assert "select.is-invalid" in css


def test_app_js_has_generic_required_form_validation() -> None:
    script = Path("app/static/app.js").read_text(encoding="utf-8")

    assert "formRequiredFieldsReady" in script
    assert "markInvalidControls" in script
    assert "focusFirstInvalidControl" in script
    assert "data-product-edit-submit" in script


def test_dashboard_keeps_primary_navigation_cards(ui_client: TestClient) -> None:
    response = ui_client.get("/")

    assert response.status_code == 200
    assert 'href="/upload"' in response.text
    assert 'href="/generate"' in response.text
    assert 'href="/search"' in response.text
    assert 'href="/data-quality"' in response.text
    assert 'href="/hs-codes/upload"' in response.text
    assert 'href="/hs-codes/generate"' in response.text


def test_upload_preview_loading_has_streaming_table_controls(ui_client: TestClient) -> None:
    response = ui_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "ui-upload.xlsx",
                build_quotation_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert 'class="preview-page-form"' in response.text
    assert "data-upload-preview-stream=" in response.text
    assert "data-upload-preview-body" in response.text
    assert "data-upload-preview-confirm disabled" in response.text
    assert "data-upload-preview-confirm-spinner" in response.text
    assert 'class="upload-preview-table"' in response.text


def test_search_page_groups_product_and_history_without_repeated_edit_links(
    ui_client: TestClient,
) -> None:
    insert_search_fixture(ui_client.database_path)

    response = ui_client.get("/search", params={"field": "gts_no", "q": "GTSUI001"})

    assert response.status_code == 200
    assert response.text.count('class="product-result"') == 1
    assert response.text.count('href="/products/1/edit"') == 1
    assert "历史报价 2 条" in response.text
    assert "Factory A" in response.text
    assert "Factory B" in response.text


def test_data_quality_page_keeps_four_spaced_sections(ui_client: TestClient) -> None:
    response = ui_client.get("/data-quality")

    assert response.status_code == 200
    assert 'class="summary-grid data-quality-summary"' in response.text
    assert response.text.count('class="table-section data-quality-section"') == 4
    assert "缺少 HS Code" in response.text
    assert "缺少 OEM" in response.text
    assert "缺少英文描述" in response.text
    assert "没有历史报价" in response.text


def build_quotation_workbook() -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    headers = [
        "No.",
        "GTS No.",
        "Description",
        "OEM",
        "Photo",
        "Factory",
        "Chinese Description",
        "Quantity",
        "Unit",
        "Unit Price",
    ]
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=3, column=column_index, value=header)
    values = [1, "GTS-UI-001", "Mirror", "OEM-UI-001", "", "Factory A", "镜子", 1, "pc", 10]
    for column_index, value in enumerate(values, start=1):
        worksheet.cell(row=4, column=column_index, value=value)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def insert_search_fixture(database_path: Path) -> None:
    connection = sqlite3.connect(database_path)
    try:
        connection.execute(
            """
            INSERT INTO products (
                id, gts_no, gts_no_normalized, oem, oem_normalized,
                description, chinese_description, hs_code,
                created_by, created_at, updated_by, updated_at
            )
            VALUES (
                1, 'GTS-UI-001', 'GTSUI001', 'OEM-UI-001', 'OEMUI001',
                'Mirror', '镜子', '87089910',
                'Tester', '2026-01-01T00:00:00+00:00', 'Tester', '2026-01-01T00:00:00+00:00'
            )
            """
        )
        connection.execute(
            """
            INSERT INTO quotation_items (
                id, product_id, factory, unit_price,
                created_by, created_at, updated_by, updated_at
            )
            VALUES (
                1, 1, 'Factory A', 10,
                'Tester', '2026-01-01T00:00:00+00:00', 'Tester', '2026-01-01T00:00:00+00:00'
            )
            """
        )
        connection.execute(
            """
            INSERT INTO quotation_items (
                id, product_id, factory, unit_price,
                created_by, created_at, updated_by, updated_at
            )
            VALUES (
                2, 1, 'Factory B', 12,
                'Tester', '2026-01-02T00:00:00+00:00', 'Tester', '2026-01-02T00:00:00+00:00'
            )
            """
        )
        connection.commit()
    finally:
        connection.close()
