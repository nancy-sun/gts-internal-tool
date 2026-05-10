from pathlib import Path

from openpyxl import Workbook

from app.services.request_parser import parse_request_workbook


def test_parse_request_workbook_accepts_only_gts_column(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "GTS No."
    sheet["A2"] = "GTS-100"
    path = tmp_path / "request_gts_only.xlsx"
    workbook.save(path)

    rows = parse_request_workbook(path)

    assert len(rows) == 1
    assert rows[0].values["gts_no_normalized"] == "GTS100"
    assert rows[0].values["description"] == ""
    assert rows[0].values["oem"] == ""
    assert rows[0].values["quantity"] is None
    assert rows[0].warnings == ["未填写数量"]


def test_parse_request_workbook_accepts_gts_and_description_with_extra_columns(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A3"] = "Random"
    sheet["B3"] = "GTS"
    sheet["C3"] = "Extra"
    sheet["D3"] = "Desc."
    sheet["E3"] = "Qty"
    sheet["F3"] = "Unit"
    sheet["B4"] = "GTS-200"
    sheet["D4"] = "Uploaded Description"
    sheet["E4"] = 5
    sheet["F4"] = "SET"
    path = tmp_path / "request_gts_description.xlsx"
    workbook.save(path)

    rows = parse_request_workbook(path)

    assert len(rows) == 1
    assert rows[0].values["gts_no_normalized"] == "GTS200"
    assert rows[0].values["description"] == "Uploaded Description"
    assert rows[0].values["quantity"] == 5
    assert rows[0].values["unit"] == "SET"


def test_parse_request_workbook_uses_aliases_before_fallback_columns(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A3"] = "Random"
    sheet["B3"] = "Wrong GTS Header"
    sheet["C3"] = "Desc"
    sheet["D3"] = "OEM"
    sheet["E3"] = "GTS"
    sheet["G3"] = "Quantity"
    sheet["B4"] = "GTS-FALLBACK"
    sheet["C4"] = "Fallback Description"
    sheet["D4"] = "OEM-FALLBACK"
    sheet["E4"] = "GTS-ALIAS"
    sheet["G4"] = 4
    path = tmp_path / "request_alias_before_fallback.xlsx"
    workbook.save(path)

    rows = parse_request_workbook(path)

    assert len(rows) == 1
    assert rows[0].values["gts_no"] == "GTS-ALIAS"
    assert rows[0].values["description"] == "Fallback Description"
    assert rows[0].values["oem"] == "OEM-FALLBACK"
    assert rows[0].values["quantity"] == 4
