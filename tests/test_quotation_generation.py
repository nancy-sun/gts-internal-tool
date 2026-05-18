import sqlite3

from openpyxl import load_workbook

from app.services.operation_logging import utc_now_text
from app.services.quotation_generation import (
    GENERATED_COLUMNS,
    apply_generated_workbook_formatting,
    build_generation_preview,
    build_output_row,
    create_generated_workbook,
)
from app.services.request_parser import ParsedRequestRow


class CountingConnection:
    def __init__(self, connection: sqlite3.Connection) -> None:
        self.connection = connection
        self.select_count = 0

    def execute(self, sql: str, parameters=()):
        if sql.lstrip().upper().startswith("SELECT"):
            self.select_count += 1
        return self.connection.execute(sql, parameters)


def test_generated_columns_include_required_updated_fields():
    labels = [label for _, label in GENERATED_COLUMNS]
    assert labels[-2:] == ["Updated By", "Updated At"]
    assert "Total Price" in labels
    assert labels[4] == "Photo"


def test_build_output_row_recalculates_total_price():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO quotation_items (
            id, product_id, unit_price, total_price, updated_by, updated_at,
            created_by, created_at
        )
        VALUES (1, 1, 12.5, 999, 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    candidate = connection.execute("SELECT * FROM quotation_items WHERE id = 1").fetchone()

    output = build_output_row(candidate, None, 4)

    assert output["quantity"] == 4
    assert output["total_price"] == 50
    assert output["updated_at"] == now[:10]


def test_build_output_row_uses_whole_number_quantity_for_export_and_total():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO quotation_items (
            id, product_id, unit_price, total_price, updated_by, updated_at,
            created_by, created_at
        )
        VALUES (1, 1, 10, 999, 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    candidate = connection.execute("SELECT * FROM quotation_items WHERE id = 1").fetchone()

    output = build_output_row(candidate, None, 3.4)

    assert output["quantity"] == 3
    assert output["total_price"] == 30


def test_build_output_row_leaves_total_price_blank_without_request_quantity():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO quotation_items (
            id, product_id, unit_price, total_price, updated_by, updated_at,
            created_by, created_at
        )
        VALUES (1, 1, 12.5, 999, 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    candidate = connection.execute("SELECT * FROM quotation_items WHERE id = 1").fetchone()

    output = build_output_row(candidate, None, None)

    assert output["quantity"] is None
    assert output["total_price"] is None


def test_build_output_row_uses_current_product_oem_description_and_request_unit():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO quotation_items (
            id, product_id, description, oem, chinese_description, unit, updated_by, updated_at,
            created_by, created_at
        )
        VALUES (1, 1, 'Historical Description', 'HIST-OEM', '历史描述', 'PCS', 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    candidate = connection.execute("SELECT * FROM quotation_items WHERE id = 1").fetchone()
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, description, oem, oem_normalized,
            chinese_description, created_by, created_at, updated_by, updated_at
        )
        VALUES (1, 'GTS-CURRENT', 'GTSCURRENT', 'Current Product Description',
                'CURRENT-OEM', 'CURRENTOEM', '当前品名', 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    product = connection.execute("SELECT * FROM products WHERE id = 1").fetchone()

    output = build_output_row(
        candidate,
        product,
        None,
        "Uploaded Request Description",
        "REQ-OEM",
        "SET",
    )

    assert output["description"] == "Current Product Description"
    assert output["oem"] == "CURRENT-OEM"
    assert output["unit"] == "SET"
    assert output["chinese_description"] == "当前品名"


def test_build_output_row_uses_system_oem_description_and_unit_when_request_values_empty():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO quotation_items (
            id, product_id, description, oem, unit, updated_by, updated_at,
            created_by, created_at
        )
        VALUES (1, 1, 'Historical Description', 'HIST-OEM', 'PCS', 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    candidate = connection.execute("SELECT * FROM quotation_items WHERE id = 1").fetchone()

    output = build_output_row(candidate, None, None, "", "", "")

    assert output["description"] == "Historical Description"
    assert output["oem"] == "HIST-OEM"
    assert output["unit"] == "PCS"


def test_apply_generated_workbook_formatting_sets_requested_number_formats():
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    for index, (_, label) in enumerate(GENERATED_COLUMNS, start=1):
        worksheet.cell(row=2, column=index, value=label)
    worksheet["H3"] = 4
    worksheet["J3"] = 12.5
    worksheet["K3"] = 50
    worksheet["P3"] = 1.234
    worksheet["Q3"] = 2.345
    worksheet["R3"] = 3.456
    worksheet["S3"] = 0.789

    apply_generated_workbook_formatting(worksheet)

    assert worksheet["H3"].number_format == "0"
    assert worksheet["J3"].number_format == '"¥"#,##0.00'
    assert worksheet["K3"].number_format == '"¥"#,##0.00'
    assert worksheet["P3"].number_format == "0.00"
    assert worksheet["Q3"].number_format == "0.00"
    assert worksheet["R3"].number_format == "0.00"
    assert worksheet["S3"].number_format == "0.00"


def test_build_generation_preview_marks_multiple_candidates():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, created_by, created_at, updated_by, updated_at
        )
        VALUES (1, 'GTS-1', 'GTS1', 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    for item_id in (1, 2):
        connection.execute(
            """
            INSERT INTO quotation_items (
                id, product_id, unit_price, updated_by, updated_at, created_by, created_at
            )
            VALUES (?, 1, ?, 'Alice', ?, 'Alice', ?)
            """,
            (item_id, item_id * 10, now, now),
        )
    request_row = ParsedRequestRow(
        row_number=4,
        values={"gts_no_normalized": "GTS1", "oem_normalized": "", "gts_no": "GTS-1", "oem": "", "quantity": 1, "comment": ""},
        warnings=[],
        errors=[],
    )

    preview = build_generation_preview(connection, [request_row])

    assert preview[0]["status"] == "multiple_candidates"
    assert preview[0]["candidates"][0]["id"] == 2


def test_build_generation_preview_dedupes_same_gts_factory_unit_and_price_candidates():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, created_by, created_at, updated_by, updated_at
        )
        VALUES (1, 'GTSTEST014', 'GTSTEST014', 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    connection.execute(
        """
        INSERT INTO quotation_items (
            id, product_id, gts_no, gts_no_normalized, oem, oem_normalized,
            factory, unit, unit_price, updated_by, updated_at, created_by, created_at
        )
        VALUES
            (1, 1, 'GTSTEST014', 'GTSTEST014', '7482541385', '7482541385',
             '鼎佳', 'pc', 600, 'Nancy Sun', '2026-05-09T17:25:33+00:00', 'Nancy Sun', ?),
            (2, 1, 'GTSTEST014', 'GTSTEST014', '1482541385', '1482541385',
             '鼎佳', 'pc', 600, 'Nancy Sun', '2026-05-09T19:50:06+00:00', 'Nancy Sun', ?)
        """,
        (now, now),
    )
    request_row = ParsedRequestRow(
        row_number=17,
        values={
            "gts_no_normalized": "GTSTEST014",
            "oem_normalized": "",
            "gts_no": "GTSTEST014",
            "oem": "",
            "quantity": 10,
            "unit": "",
            "comment": "",
        },
        warnings=[],
        errors=[],
    )

    preview = build_generation_preview(connection, [request_row])

    assert preview[0]["status"] == "ready"
    assert len(preview[0]["candidates"]) == 1
    assert preview[0]["candidates"][0]["id"] == 2


def test_build_generation_preview_uses_batch_lookups_for_repeated_products():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, created_by, created_at, updated_by, updated_at
        )
        VALUES (1, 'GTS0001', 'GTS0001', 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    connection.execute(
        """
        INSERT INTO quotation_items (
            id, product_id, gts_no, gts_no_normalized, factory, unit, unit_price,
            updated_by, updated_at, created_by, created_at
        )
        VALUES (1, 1, 'GTS0001', 'GTS0001', 'Factory A', 'pc', 20,
                'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    parsed_rows = [
        ParsedRequestRow(
            row_number=row_number,
            values={
                "gts_no_normalized": "GTS0001",
                "oem_normalized": "",
                "gts_no": "GTS0001",
                "oem": "",
                "quantity": 1,
                "unit": "",
                "comment": "",
            },
            warnings=[],
            errors=[],
        )
        for row_number in (4, 5, 6)
    ]
    counting_connection = CountingConnection(connection)

    preview = build_generation_preview(counting_connection, parsed_rows)

    assert [row["product"]["id"] for row in preview] == [1, 1, 1]
    assert [row["candidates"][0]["id"] for row in preview] == [1, 1, 1]
    assert counting_connection.select_count == 3


def test_build_generation_preview_matches_by_oem_when_gts_is_missing():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, oem, oem_normalized,
            created_by, created_at, updated_by, updated_at
        )
        VALUES (1, 'GTS-1', 'GTS1', 'OEM-1', 'OEM1', 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    connection.execute(
        """
        INSERT INTO quotation_items (
            id, product_id, unit_price, updated_by, updated_at, created_by, created_at
        )
        VALUES (1, 1, 20, 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    request_row = ParsedRequestRow(
        row_number=4,
        values={"gts_no_normalized": "", "oem_normalized": "OEM1", "gts_no": "", "oem": "OEM-1", "quantity": 1, "unit": "", "comment": ""},
        warnings=[],
        errors=[],
    )

    preview = build_generation_preview(connection, [request_row])

    assert preview[0]["status"] == "ready"
    assert preview[0]["product"]["id"] == 1
    assert preview[0]["candidates"][0]["unit_price"] == 20


def test_build_generation_preview_marks_unmatched_and_no_quotation_rows():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, created_by, created_at, updated_by, updated_at
        )
        VALUES (1, 'GTS-NOQUOTE', 'GTSNOQUOTE', 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    rows = [
        ParsedRequestRow(
            row_number=4,
            values={"gts_no_normalized": "MISSING", "oem_normalized": "", "gts_no": "MISSING", "oem": "", "quantity": 1, "unit": "", "comment": ""},
            warnings=[],
            errors=[],
        ),
        ParsedRequestRow(
            row_number=5,
            values={"gts_no_normalized": "GTSNOQUOTE", "oem_normalized": "", "gts_no": "GTS-NOQUOTE", "oem": "", "quantity": 1, "unit": "", "comment": ""},
            warnings=[],
            errors=[],
        ),
    ]

    preview = build_generation_preview(connection, rows)

    assert [row["status"] for row in preview] == ["unmatched", "no_quotation"]


def test_build_generation_preview_marks_gts_oem_conflict():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, oem, oem_normalized,
            created_by, created_at, updated_by, updated_at
        )
        VALUES (1, 'GTS-1', 'GTS1', 'OEM-1', 'OEM1', 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, oem, oem_normalized,
            created_by, created_at, updated_by, updated_at
        )
        VALUES (2, 'GTS-2', 'GTS2', 'OEM-2', 'OEM2', 'Alice', ?, 'Alice', ?)
        """,
        (now, now),
    )
    request_row = ParsedRequestRow(
        row_number=4,
        values={"gts_no_normalized": "GTS1", "oem_normalized": "OEM2", "gts_no": "GTS-1", "oem": "OEM-2", "quantity": 1, "unit": "", "comment": ""},
        warnings=[],
        errors=[],
    )

    preview = build_generation_preview(connection, [request_row])

    assert preview[0]["status"] == "conflict"
    assert preview[0]["errors"] == ["GTS 和 OEM 匹配到不同产品，需要人工确认。"]


def test_create_generated_workbook_excludes_unchecked_rows_and_keeps_blank_photo_column():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    initialize_test_schema(connection)
    now = utc_now_text()
    connection.execute(
        """
        INSERT INTO products (
            id, gts_no, gts_no_normalized, description, oem, oem_normalized,
            created_by, created_at, updated_by, updated_at
        )
        VALUES
            (1, 'GTS-CURRENT-1', 'GTSCURRENT1', 'Current 1', 'OEM-CURRENT-1',
             'OEMCURRENT1', 'Alice', ?, 'Alice', ?),
            (2, 'GTS-CURRENT-2', 'GTSCURRENT2', 'Current 2', 'OEM-CURRENT-2',
             'OEMCURRENT2', 'Alice', ?, 'Alice', ?)
        """,
        (now, now, now, now),
    )
    connection.execute(
        """
        INSERT INTO quotation_items (
            id, product_id, gts_no, description, oem, unit, unit_price, updated_by, updated_at,
            created_by, created_at
        )
        VALUES
            (1, 1, 'GTS-1', 'Historical 1', 'OEM-1', 'PCS', 10, 'Alice', ?, 'Alice', ?),
            (2, 2, 'GTS-2', 'Historical 2', 'OEM-2', 'PCS', 20, 'Alice', ?, 'Alice', ?)
        """,
        (now, now, now, now),
    )
    preview_rows = [
        {
            "row_number": 4,
            "values": {"quantity": 3, "description": "", "oem": "", "unit": ""},
        },
        {
            "row_number": 5,
            "values": {"quantity": 4, "description": "", "oem": "", "unit": ""},
        },
    ]

    stream, generated_count = create_generated_workbook(
        connection,
        preview_rows=preview_rows,
        selected_candidate_ids={4: 1},
        operator_name="Nancy",
        request_file_name="request.xlsx",
    )
    workbook = load_workbook(stream, data_only=True)
    worksheet = workbook.active

    assert generated_count == 1
    assert worksheet["A3"].value == 1
    assert worksheet["B3"].value == "GTS-CURRENT-1"
    assert worksheet["E3"].value is None
    assert worksheet["K3"].value == 30
    assert worksheet["X3"].value == now[:10]
    assert worksheet.max_row == 3
    log = connection.execute("SELECT * FROM operation_logs").fetchone()
    assert log["action_type"] == "generate_quotation"
    assert log["row_count"] == 1


def initialize_test_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        CREATE TABLE products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            gts_no TEXT,
            gts_no_normalized TEXT,
            oem TEXT,
            oem_normalized TEXT,
            description TEXT,
            chinese_description TEXT,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_by TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE quotation_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            no TEXT,
            gts_no TEXT,
            gts_no_normalized TEXT,
            description TEXT,
            oem TEXT,
            oem_normalized TEXT,
            factory TEXT,
            chinese_description TEXT,
            quantity REAL,
            unit TEXT,
            unit_price REAL,
            total_price REAL,
            item_per_package TEXT,
            packages TEXT,
            weight_per_package TEXT,
            gross_weight TEXT,
            length TEXT,
            width TEXT,
            height TEXT,
            measurements_volume TEXT,
            packaging TEXT,
            expected_delivery TEXT,
            comment TEXT,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_by TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE operation_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action_time TEXT NOT NULL,
            operator_name TEXT NOT NULL,
            action_type TEXT NOT NULL,
            file_name TEXT,
            row_count INTEGER,
            note TEXT
        );
        """
    )
