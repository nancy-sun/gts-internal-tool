from types import SimpleNamespace
from io import BytesIO

from openpyxl import Workbook

from app.services.upload_validation import (
    sanitize_upload_filename,
    validate_full_quotation_workbook,
    validate_request_list_workbook,
    validate_upload_size,
    validate_workbook_contents,
    validate_xlsx_upload,
)


def test_validate_xlsx_upload_requires_operator_name():
    upload_file = SimpleNamespace(filename="quotation.xlsx")

    assert validate_xlsx_upload(upload_file, " ") == "请填写操作员姓名。"


def test_validate_xlsx_upload_allows_only_xlsx_files():
    upload_file = SimpleNamespace(filename="quotation.xls")

    assert validate_xlsx_upload(upload_file, "Nancy") == "只能上传 .xlsx 文件。"


def test_validate_workbook_contents_rejects_corrupted_xlsx():
    message = validate_workbook_contents(b"not a real workbook")

    assert message == "Excel 文件无法读取，请确认文件是有效的 .xlsx 工作簿。"


def test_validate_full_quotation_workbook_rejects_missing_expected_header():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Random"
    sheet["B1"] = "Other"

    message = validate_full_quotation_workbook(workbook_bytes(workbook))

    assert message == "系统找不到报价单表头行，请确认 Excel 中有 No.、GTS、OEM、工厂、单位、价格等表头。"


def test_validate_full_quotation_workbook_accepts_expected_header():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A3"] = "No."
    sheet["B3"] = "GTS No."
    sheet["D3"] = "OEM"
    sheet["F3"] = "Factory"
    sheet["I3"] = "Unit"
    sheet["J3"] = "Unit Price"

    assert validate_full_quotation_workbook(workbook_bytes(workbook)) is None


def test_validate_request_list_workbook_requires_gts_or_oem_header():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Description"
    sheet["B1"] = "Quantity"

    message = validate_request_list_workbook(workbook_bytes(workbook))

    assert message == "需求文件中找不到 GTS 或 OEM 表头，请检查需求 Excel。"


def test_validate_request_list_workbook_accepts_gts_header():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "GTS"
    sheet["A2"] = "GTS-001"

    assert validate_request_list_workbook(workbook_bytes(workbook)) is None


def test_sanitize_upload_filename_removes_paths_limits_length_and_keeps_extension():
    filename = "../folder\\subfolder/" + ("quotation-" * 30) + ".xlsx"

    safe_name = sanitize_upload_filename(filename, default="quotation.xlsx", max_length=40)

    assert "/" not in safe_name
    assert "\\" not in safe_name
    assert len(safe_name) <= 40
    assert safe_name.endswith(".xlsx")


def test_validate_upload_size_reports_limit():
    message = validate_upload_size(
        b"1234",
        max_upload_size_bytes=3,
        max_upload_size_mb=10,
    )

    assert message == "上传文件超过 10 MB。"


def workbook_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
