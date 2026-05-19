from pathlib import Path

from openpyxl import Workbook

from app.services.excel_parser import parse_full_quotation_workbook


def test_parse_full_quotation_workbook_detects_header_row_and_skips_photo_column(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A6"] = "No."
    sheet["B6"] = "GTS No."
    sheet["C6"] = "Description"
    sheet["D6"] = "OEM"
    sheet["E6"] = "Photo"
    sheet["F6"] = "Factory"
    sheet["H6"] = "Quantity"
    sheet["J6"] = "Unit Price"
    sheet["A7"] = 1
    sheet["B7"] = "GTS-00123"
    sheet["C7"] = "Brake Pad"
    sheet["D7"] = " A 500 008 3949 "
    sheet["E7"] = "ignored photo cell"
    sheet["F7"] = "Factory A"
    sheet["H7"] = 2
    sheet["J7"] = 10.5
    path = tmp_path / "quotation.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(path)

    assert len(rows) == 1
    assert rows[0].values["gts_no_normalized"] == "GTS00123"
    assert rows[0].values["oem_normalized"] == "A5000083949"
    assert rows[0].values["factory"] == "Factory A"
    assert rows[0].values["quantity"] == 2
    assert rows[0].values["unit_price"] == 10.5


def test_parse_full_quotation_workbook_ignores_inserted_extra_columns(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        "No.",
        "GTS No.",
        "Description",
        "OEM",
        "Extra Column",
        "Photo",
        "Factory",
        "Chinese Description",
        "Quantity",
        "Unit",
        "Unit Price",
        "Total Price",
        "Another Extra",
        "Item/Package",
        "Packages",
        "Weight / Package",
        "G.W.",
        "Length",
        "Width",
        "Height",
        "Measurements (Volume)",
        "Packaging",
        "Expected Delivery",
        "Comment",
    ]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=index, value=header)

    sheet["A4"] = 1
    sheet["B4"] = "GTS-00999"
    sheet["C4"] = "Filter"
    sheet["D4"] = "OEM-00999"
    sheet["E4"] = "ignored extra"
    sheet["F4"] = "ignored photo"
    sheet["G4"] = "Factory Extra"
    sheet["I4"] = 3
    sheet["K4"] = 6.5
    sheet["L4"] = 19.5
    sheet["X4"] = "ok"
    path = tmp_path / "quotation_extra_columns.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(path)

    assert len(rows) == 1
    assert rows[0].values["gts_no_normalized"] == "GTS00999"
    assert rows[0].values["factory"] == "Factory Extra"
    assert rows[0].values["quantity"] == 3
    assert rows[0].values["unit_price"] == 6.5
    assert rows[0].values["total_price"] == 19.5
    assert rows[0].values["comment"] == "ok"


def test_parse_full_quotation_workbook_supports_common_header_aliases(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A2"] = "No"
    sheet["B2"] = "GTS"
    sheet["C2"] = "Desc"
    sheet["D2"] = "OEM No"
    sheet["E2"] = "Qty"
    sheet["F2"] = "Price"
    sheet["A3"] = 1
    sheet["B3"] = "GTS-ABC"
    sheet["C3"] = "Alias Product"
    sheet["D3"] = "OEM-ABC"
    sheet["E3"] = 8
    sheet["F3"] = 2.25
    path = tmp_path / "quotation_aliases.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(path)

    assert len(rows) == 1
    assert rows[0].values["gts_no_normalized"] == "GTSABC"
    assert rows[0].values["description"] == "Alias Product"
    assert rows[0].values["oem_normalized"] == "OEMABC"
    assert rows[0].values["quantity"] == 8
    assert rows[0].values["unit_price"] == 2.25


def test_parse_full_quotation_workbook_uses_aliases_before_fallback_columns(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A3"] = "No."
    sheet["B3"] = "Wrong GTS Header"
    sheet["C3"] = "Wrong Description Header"
    sheet["D3"] = "OEM"
    sheet["E3"] = "GTS"
    sheet["F3"] = "Factory"
    sheet["H3"] = "Quantity"
    sheet["J3"] = "Price"
    sheet["A4"] = 1
    sheet["B4"] = "GTS-FALLBACK"
    sheet["C4"] = "Fallback Description"
    sheet["D4"] = "OEM-FALLBACK"
    sheet["E4"] = "GTS-ALIAS"
    sheet["F4"] = "Factory Fallback"
    sheet["H4"] = 3
    sheet["J4"] = 7.5
    path = tmp_path / "quotation_alias_before_fallback.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(path)

    assert len(rows) == 1
    assert rows[0].values["gts_no"] == "GTS-ALIAS"
    assert rows[0].values["description"] == "Fallback Description"
    assert rows[0].values["oem"] == "OEM-FALLBACK"
    assert rows[0].values["factory"] == "Factory Fallback"
    assert rows[0].values["quantity"] == 3
    assert rows[0].values["unit_price"] == 7.5


def test_parse_full_quotation_workbook_skips_repeated_header_rows(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    headers = ["No.", "GTS No.", "Description", "OEM", "Factory", "Quantity", "Unit Price"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=index, value=header)
        sheet.cell(row=5, column=index, value=header)

    sheet["A4"] = 1
    sheet["B4"] = "GTS-001"
    sheet["C4"] = "First Product"
    sheet["D4"] = "OEM-001"
    sheet["E4"] = "Factory A"
    sheet["F4"] = 2
    sheet["G4"] = 10
    sheet["A6"] = 2
    sheet["B6"] = "GTS-002"
    sheet["C6"] = "Second Product"
    sheet["D6"] = "OEM-002"
    sheet["E6"] = "Factory B"
    sheet["F6"] = 3
    sheet["G6"] = 12
    path = tmp_path / "quotation_repeated_header.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(path)

    assert [row.row_number for row in rows] == [4, 6]
    assert [row.values["gts_no"] for row in rows] == ["GTS-001", "GTS-002"]


def test_parse_full_quotation_workbook_detects_alias_header_without_no_column(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["B4"] = "品名"
    sheet["C4"] = "oem"
    sheet["D4"] = "数量"
    sheet["E4"] = "图片"
    sheet["F4"] = "价格"
    sheet["B5"] = "MIRROR GLASS BIG RH=LH"
    sheet["C5"] = "7420862795"
    sheet["D5"] = 60
    sheet["E5"] = "ignored image"
    sheet["F5"] = 9.5
    path = tmp_path / "quotation_alias_header_without_no.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(path)

    assert len(rows) == 1
    assert rows[0].row_number == 5
    assert rows[0].values["gts_no"] == ""
    assert rows[0].values["description"] == ""
    assert rows[0].values["chinese_description"] == "MIRROR GLASS BIG RH=LH"
    assert rows[0].values["oem"] == "7420862795"
    assert rows[0].values["quantity"] == 60
    assert rows[0].values["unit_price"] == 9.5


def test_parse_full_quotation_workbook_supports_extended_case_insensitive_aliases(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        "no",
        "gts",
        "desc.",
        "oem.",
        "工厂",
        "品名",
        "qty",
        "单位",
        "prix",
        "amount",
        "item/pkg",
        "pkg.",
        "w./pkg",
        "毛重",
        "l.",
        "w.",
        "h.",
        "vol.",
        "包装",
        "delivery date",
        "备注",
    ]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=5, column=index, value=header)

    values = [
        1,
        "gts-lower",
        "Lowercase Alias Product",
        "oem-lower",
        "工厂A",
        "中文产品",
        6,
        "PCS",
        3.5,
        21,
        "12/CTN",
        2,
        "5 kg",
        "10 kg",
        10,
        20,
        30,
        "0.06 CBM",
        "纸箱",
        "45 days",
        "测试备注",
    ]
    for index, value in enumerate(values, start=1):
        sheet.cell(row=6, column=index, value=value)
    path = tmp_path / "quotation_extended_aliases.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(path)

    assert len(rows) == 1
    assert rows[0].values["gts_no_normalized"] == "GTSLOWER"
    assert rows[0].values["description"] == "Lowercase Alias Product"
    assert rows[0].values["oem_normalized"] == "OEMLOWER"
    assert rows[0].values["factory"] == "工厂A"
    assert rows[0].values["chinese_description"] == "中文产品"
    assert rows[0].values["quantity"] == 6
    assert rows[0].values["unit"] == "PCS"
    assert rows[0].values["unit_price"] == 3.5
    assert rows[0].values["total_price"] == 21
    assert rows[0].values["item_per_package"] == 12
    assert rows[0].values["packages"] == 2
    assert rows[0].values["weight_per_package"] == 5
    assert rows[0].values["gross_weight"] == 10
    assert rows[0].values["length"] == 10
    assert rows[0].values["width"] == 20
    assert rows[0].values["height"] == 30
    assert rows[0].values["measurements_volume"] == 0.06
    assert rows[0].values["packaging"] == "纸箱"
    assert rows[0].values["expected_delivery"] == "45 days"
    assert rows[0].values["comment"] == "测试备注"


def test_parse_full_quotation_workbook_supports_ctn_aliases_from_real_files(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        "no",
        "gts No",
        "DES",
        "OEM. No",
        "Image",
        "Factory",
        "qty",
        "unit",
        "price",
        "总价",
        "item/ctn",
        "ctn",
        "gw/ctn",
        "gw",
    ]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=index, value=header)
    values = [
        1,
        "GTSTEST001",
        "FOOT STEP RH",
        5010225393,
        "",
        "hi",
        20,
        "pc",
        9,
        180,
        2,
        10,
        9,
        90,
    ]
    for index, value in enumerate(values, start=1):
        sheet.cell(row=4, column=index, value=value)
    path = tmp_path / "quotation_ctn_aliases.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(path)

    assert len(rows) == 1
    assert rows[0].values["description"] == "FOOT STEP RH"
    assert rows[0].values["item_per_package"] == 2
    assert rows[0].values["packages"] == 10
    assert rows[0].values["weight_per_package"] == 9
    assert rows[0].values["gross_weight"] == 90


def test_parse_full_quotation_workbook_warns_when_numeric_extra_field_has_no_number(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "No."
    sheet["B1"] = "GTS No."
    sheet["F1"] = "Factory"
    sheet["I1"] = "Unit"
    sheet["J1"] = "Unit Price"
    sheet["N1"] = "Weight / Package"
    sheet["A2"] = 1
    sheet["B2"] = "GTS-401"
    sheet["F2"] = "Factory A"
    sheet["I2"] = "pc"
    sheet["J2"] = 10
    sheet["N2"] = "heavy"
    path = tmp_path / "quotation_invalid_numeric_extra.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(path)

    assert len(rows) == 1
    assert rows[0].values["weight_per_package"] is None
    assert "Weight / Package不是数字，已留空。" in rows[0].warnings


def test_parse_full_quotation_workbook_errors_when_required_columns_missing(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "No."
    sheet["B1"] = "GTS No."
    sheet["A2"] = 1
    sheet["B2"] = "GTS-300"
    path = tmp_path / "quotation_missing_columns.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(
        path,
        {
            "sheet_name": None,
            "header_row": 1,
            "detect_header_from_column": "A",
            "header_label": "No.",
            "header_scan_rows": 100,
            "header_scan_columns": 80,
            "max_rows": 300,
            "columns": {"no": "A", "gts_no": "B"},
        },
    )

    assert len(rows) == 1
    assert "工厂不能为空。" in rows[0].errors
    assert "单位不能为空。" in rows[0].errors
    assert "单价不能为空。" in rows[0].errors


def test_parse_full_quotation_workbook_errors_when_required_values_blank(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "No."
    sheet["B1"] = "GTS No."
    sheet["F1"] = "Factory"
    sheet["I1"] = "Unit"
    sheet["J1"] = "Unit Price"
    sheet["A2"] = 1
    sheet["B2"] = "GTS-400"
    path = tmp_path / "quotation_blank_required_values.xlsx"
    workbook.save(path)

    rows = parse_full_quotation_workbook(path)

    assert len(rows) == 1
    assert "单位不能为空。" in rows[0].errors
    assert "单价不能为空。" in rows[0].errors
