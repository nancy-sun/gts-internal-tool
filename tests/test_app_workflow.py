from __future__ import annotations

import json
import re
import sqlite3
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


ACCESS_CODE = "test-access-code"


@pytest.fixture()
def app_client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> TestClient:
    database_path = tmp_path / "gts-test.sqlite3"
    upload_path = tmp_path / "uploads"
    upload_path.mkdir()

    monkeypatch.setenv("SHARED_ACCESS_CODE", ACCESS_CODE)
    monkeypatch.setenv("SESSION_SECRET_KEY", "test-session-secret-key")
    monkeypatch.setenv("PRODUCT_EDIT_PASSWORD", "55123511")
    monkeypatch.setenv("DATABASE_PATH", str(database_path))

    from app.config import get_settings
    from app.main import create_app
    from app.routes import generate, hs_codes, upload
    from app.services import backup as backup_service

    get_settings.cache_clear()
    monkeypatch.setattr(upload, "UPLOAD_DIR", upload_path)
    monkeypatch.setattr(generate, "UPLOAD_DIR", upload_path)
    monkeypatch.setattr(hs_codes, "UPLOAD_DIR", upload_path)
    monkeypatch.setattr(backup_service, "AUTO_BACKUP_DIR", upload_path / "auto-backups")

    client = TestClient(create_app())
    client.upload_path = upload_path
    login_response = client.post("/login", data={"access_code": ACCESS_CODE})
    assert login_response.status_code == 200
    return client


def test_healthz_is_public_and_maintenance_requires_login(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    database_path = tmp_path / "status.sqlite3"
    monkeypatch.setenv("SHARED_ACCESS_CODE", ACCESS_CODE)
    monkeypatch.setenv("SESSION_SECRET_KEY", "test-session-secret-key")
    monkeypatch.setenv("PRODUCT_EDIT_PASSWORD", "55123511")
    monkeypatch.setenv("DATABASE_PATH", str(database_path))

    from app.config import get_settings
    from app.main import create_app

    get_settings.cache_clear()
    client = TestClient(create_app())

    health_response = client.get("/healthz")
    assert health_response.status_code == 200
    assert health_response.json() == {"status": "ok"}

    maintenance_response = client.get("/maintenance", follow_redirects=False)
    assert maintenance_response.status_code == 303
    assert maintenance_response.headers["location"] == "/login"
    get_settings.cache_clear()


def test_maintenance_page_shows_non_secret_runtime_status(
    app_client: TestClient,
    tmp_path: Path,
) -> None:
    response = app_client.get("/maintenance")

    assert response.status_code == 200
    assert "系统状态" in response.text
    assert "运行模式" in response.text
    assert "local" in response.text
    assert "数据库路径" in response.text
    assert str(tmp_path / "gts-test.sqlite3") in response.text
    assert "数据库文件存在" in response.text
    assert "数据库大小" in response.text
    assert "Uploads 路径" in response.text
    assert "Generated 路径" in response.text
    assert "自动备份路径" in response.text
    assert "最近自动备份时间" in response.text
    assert ACCESS_CODE not in response.text
    assert "test-session-secret-key" not in response.text
    assert "55123511" not in response.text


def test_office_workflow_upload_search_generate_download_and_log(
    app_client: TestClient,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from app.services import backup as backup_service

    auto_backup_dir = tmp_path / "auto-backups"
    monkeypatch.setattr(backup_service, "AUTO_BACKUP_DIR", auto_backup_dir)

    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "quotation.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "no": 1,
                            "gts_no": "GTS-TEST-001",
                            "description": "Mirror request source",
                            "oem": "5010 225 393",
                            "factory": "欧达",
                            "chinese_description": "后视镜",
                            "quantity": 2,
                            "unit": "pc",
                            "unit_price": 99,
                            "total_price": 198,
                            "item_per_package": "12/CTN",
                            "packages": 2,
                            "weight_per_package": "5 kg",
                            "gross_weight": "10 kg",
                            "length": 10,
                            "width": 20,
                            "height": 30,
                            "measurements_volume": "0.06 CBM",
                            "expected_delivery": "25days",
                            "comment": "first upload",
                        }
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert upload_response.status_code == 200
    upload_token = extract_token(upload_response.text)
    stream_response = app_client.get(f"/upload/preview/stream/{upload_token}")
    assert stream_response.status_code == 200
    assert 'event: complete' in stream_response.text
    assert '"has_errors": false' in stream_response.text

    confirm_response = confirm_upload_after_creating_suppliers(app_client, upload_token)
    assert confirm_response.status_code == 200
    assert "新增产品" in confirm_response.text
    assert not (app_client.upload_path / f"preview_{upload_token}.json").exists()
    backup_files = list(auto_backup_dir.glob("*_full_quotation_import.sqlite3"))
    assert len(backup_files) == 1

    search_response = app_client.get("/search", params={"field": "gts_no", "q": "test001"})
    assert search_response.status_code == 200
    assert "GTS-TEST-001" in search_response.text
    assert "¥99.00" in search_response.text
    assert 'href="/products/1/edit"' in search_response.text
    assert "编辑产品" in search_response.text
    assert "历史报价 1 条" in search_response.text
    assert "只/件" in search_response.text
    assert "GW/件" in search_response.text
    assert "体积" in search_response.text
    assert ">12</td>" in search_response.text
    assert ">5</td>" in search_response.text
    assert ">0.06</td>" in search_response.text
    assert '<details class="quotation-history">' in search_response.text
    assert '<details class="quotation-history" open>' not in search_response.text

    generate_response = app_client.post(
        "/generate/preview",
        data={"operator_name": "Nancy"},
        files={
            "request_file": (
                "request.xlsx",
                build_request_workbook(
                    [
                        {
                            "gts_no": "GTSTEST001",
                            "description": "Uploaded request description",
                            "oem": "",
                            "quantity": 3,
                            "unit": "",
                            "comment": "need quote",
                        }
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert generate_response.status_code == 200
    assert_breadcrumb(
        generate_response.text,
        ["首页", "生成报价单", "生成预览"],
    )
    assert "后视镜" in generate_response.text
    assert "pc" in generate_response.text
    assert "欧达" in generate_response.text
    assert "¥99.00" in generate_response.text
    assert "Nancy" in generate_response.text
    assert ">3<" in generate_response.text
    assert "3.0" not in generate_response.text

    generate_token = extract_token(generate_response.text)
    candidate_id = fetch_single_candidate_id(tmp_path / "gts-test.sqlite3")
    download_response = app_client.post(
        "/generate/download",
        data={
            "token": generate_token,
            "include__2": "1",
            "candidate__2": str(candidate_id),
        },
    )

    assert download_response.status_code == 200
    assert "filename*=UTF-8''Nancy-%E8%AF%A2%E4%BB%B7-" in download_response.headers[
        "content-disposition"
    ]
    assert not (app_client.upload_path / f"generate_preview_{generate_token}.json").exists()
    workbook = load_workbook(BytesIO(download_response.content))
    worksheet = workbook.active
    assert [worksheet.cell(row=2, column=index).value for index in range(1, 6)] == [
        "No.",
        "GTS No.",
        "Description",
        "OEM",
        "Photo",
    ]
    assert worksheet["B3"].value == "GTS-TEST-001"
    assert worksheet["C3"].value == "Mirror request source"
    assert worksheet["D3"].value == "5010 225 393"
    assert worksheet["G3"].value == "后视镜"
    assert worksheet["H3"].value == 3
    assert worksheet["I3"].value == "pc"
    assert worksheet["J3"].value == 99
    assert worksheet["K3"].value == 297

    logs_response = app_client.get("/logs")
    assert logs_response.status_code == 200
    assert_breadcrumb(logs_response.text, ["首页", "操作记录"])
    assert "上传完整报价单" in logs_response.text
    assert "生成报价单" in logs_response.text
    assert str(backup_files[0]) in logs_response.text


def test_page_breadcrumbs_include_parent_pages(app_client: TestClient) -> None:
    upload_page = app_client.get("/upload")
    assert upload_page.status_code == 200
    assert_breadcrumb(upload_page.text, ["首页", "上传完整报价单"])

    upload_preview = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "breadcrumb-upload.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "gts_no": "GTS-BREAD-001",
                            "oem": "OEM-BREAD",
                            "factory": "Factory A",
                            "unit": "pc",
                            "unit_price": 10,
                        }
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_preview.status_code == 200
    assert_breadcrumb(upload_preview.text, ["首页", "上传完整报价单", "导入预览"])

    upload_token = extract_token(upload_preview.text)
    stream_response = app_client.get(f"/upload/preview/stream/{upload_token}")
    assert stream_response.status_code == 200
    upload_result = confirm_upload_after_creating_suppliers(app_client, upload_token)
    assert upload_result.status_code == 200
    assert_breadcrumb(upload_result.text, ["首页", "上传完整报价单", "导入结果"])

    generate_page = app_client.get("/generate")
    assert generate_page.status_code == 200
    assert_breadcrumb(generate_page.text, ["首页", "生成报价单"])

    generate_preview = app_client.post(
        "/generate/preview",
        data={"operator_name": "Nancy"},
        files={
            "request_file": (
                "breadcrumb-request.xlsx",
                build_request_workbook([{"gts_no": "GTS-BREAD-001", "quantity": 1}]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert generate_preview.status_code == 200
    assert_breadcrumb(generate_preview.text, ["首页", "生成报价单", "生成预览"])

    search_page = app_client.get("/search")
    assert search_page.status_code == 200
    assert_breadcrumb(search_page.text, ["首页", "查询数据库"])

    data_quality_page = app_client.get("/data-quality")
    assert data_quality_page.status_code == 200
    assert_breadcrumb(data_quality_page.text, ["首页", "数据检查"])
    assert "缺少 HS Code" in data_quality_page.text
    assert "没有历史报价" in data_quality_page.text

    logs_page = app_client.get("/logs")
    assert logs_page.status_code == 200
    assert_breadcrumb(logs_page.text, ["首页", "操作记录"])


def test_upload_preview_stream_replaces_loading_rows_one_by_one(app_client: TestClient) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "stream-order.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "gts_no": "GTS-STREAM-001",
                            "oem": "OEM-STREAM-001",
                            "factory": "Factory A",
                            "unit": "pc",
                            "unit_price": 10,
                        },
                        {
                            "gts_no": "GTS-STREAM-002",
                            "oem": "OEM-STREAM-002",
                            "factory": "Factory B",
                            "unit": "pc",
                            "unit_price": 20,
                        },
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 200

    token = extract_token(upload_response.text)
    stream_response = app_client.get(f"/upload/preview/stream/{token}")

    assert stream_response.status_code == 200
    first_loading_index = stream_response.text.index(
        'event: loading\ndata: {"row_number": 4}'
    )
    first_row_index = stream_response.text.index(
        'event: row\ndata: {"row_number": 4'
    )
    second_loading_index = stream_response.text.index(
        'event: loading\ndata: {"row_number": 5}'
    )
    assert first_loading_index < first_row_index < second_loading_index


def test_upload_preview_stream_returns_friendly_error_for_bad_workbook(
    app_client: TestClient,
) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "bad.xlsx",
                BytesIO(b"not a real workbook"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 400
    assert "Excel 文件无法读取，请确认文件是有效的 .xlsx 工作簿。" in upload_response.text
    assert "File is not a zip file" not in upload_response.text


def test_upload_preview_reports_missing_required_fields(app_client: TestClient) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "missing-required.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "no": 1,
                            "gts_no": "GTS-ERR-001",
                            "oem": "OEM-ERR",
                            "factory": "",
                            "unit": "",
                            "unit_price": "",
                        }
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    token = extract_token(upload_response.text)
    stream_response = app_client.get(f"/upload/preview/stream/{token}")
    assert stream_response.status_code == 200
    assert '"has_errors": true' in stream_response.text
    assert "单位不能为空" in stream_response.text
    assert "单价不能为空" in stream_response.text

    confirm_response = app_client.post("/upload/confirm", data={"token": token})
    assert confirm_response.status_code == 400
    assert "预览有错误，请修改 Excel 后再导入。" in confirm_response.text


def test_upload_preview_groups_repeated_factory_and_blocks_until_resolved(
    app_client: TestClient,
    tmp_path: Path,
) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "repeated-factory.xlsx",
                build_quotation_workbook(
                    [
                        {"gts_no": "GTS-REP-001", "factory": "Same Factory", "unit": "pc", "unit_price": 10},
                        {"gts_no": "GTS-REP-002", "factory": "Same Factory", "unit": "pc", "unit_price": 20},
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    token = extract_token(upload_response.text)
    stream_response = app_client.get(f"/upload/preview/stream/{token}")
    payload = preview_payload(app_client, token)

    assert stream_response.status_code == 200
    assert payload["supplier_matches"][0]["status"] == "pending_unmatched"
    assert payload["supplier_matches"][0]["occurrence_count"] == 2
    assert len(payload["supplier_matches"]) == 1

    blocked_response = app_client.post("/upload/confirm", data={"token": token})
    assert blocked_response.status_code == 400
    assert "还有未处理的供应商" in blocked_response.text

    confirm_response = confirm_upload_after_creating_suppliers(app_client, token)
    assert confirm_response.status_code == 200
    with sqlite3.connect(tmp_path / "gts-test.sqlite3") as connection:
        supplier_ids = [
            row[0]
            for row in connection.execute(
                "SELECT supplier_id FROM quotation_items ORDER BY gts_no"
            ).fetchall()
        ]
    assert supplier_ids[0] == supplier_ids[1]
    assert supplier_ids[0] is not None


def test_upload_preview_uses_autocomplete_for_existing_supplier_link(
    app_client: TestClient,
) -> None:
    create_supplier_for_test(
        app_client,
        full_name="Known Supplier Full",
        short_name="Known Supplier",
        aliases_text="Known Alias",
    )
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "unmatched-factory.xlsx",
                build_quotation_workbook(
                    [{"gts_no": "GTS-AUTO-001", "factory": "New Factory", "unit": "pc", "unit_price": 10}]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    token = extract_token(upload_response.text)
    app_client.get(f"/upload/preview/stream/{token}")

    preview_page = app_client.get(f"/upload/preview/{token}")

    assert preview_page.status_code == 200
    assert 'data-supplier-combobox' in preview_page.text
    assert 'data-supplier-existing-search' in preview_page.text
    assert 'data-supplier-existing-field' in preview_page.text
    assert 'data-supplier-options' in preview_page.text
    assert 'data-supplier-option' in preview_page.text
    assert '没有匹配的供应商' in preview_page.text
    assert "Known Supplier Full" in preview_page.text
    assert "Known Supplier" in preview_page.text
    assert "Known Alias" in preview_page.text
    assert "<datalist" not in preview_page.text
    assert 'list="supplier-options-' not in preview_page.text
    assert 'select name="supplier_id__' not in preview_page.text


def test_batch_create_same_short_name_merges_multiple_factory_groups(
    app_client: TestClient,
    tmp_path: Path,
) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "same-supplier-different-factory.xlsx",
                build_quotation_workbook(
                    [
                        {"gts_no": "GTS-MERGE-001", "factory": "Factory Alias A", "unit": "pc", "unit_price": 10},
                        {"gts_no": "GTS-MERGE-002", "factory": "Factory Alias B", "unit": "pc", "unit_price": 20},
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    token = extract_token(upload_response.text)
    app_client.get(f"/upload/preview/stream/{token}")
    payload = preview_payload(app_client, token)
    matches = payload["supplier_matches"]

    assert len(matches) == 2
    response = app_client.post(
        "/upload/preview/supplier/batch",
        data={
            "token": token,
            "operator_name": "Nancy",
            f"action__{matches[0]['key']}": "create",
            f"supplier_short_name__{matches[0]['key']}": "Merged Supplier",
            f"action__{matches[1]['key']}": "create",
            f"supplier_short_name__{matches[1]['key']}": "Merged Supplier",
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    resolved_payload = preview_payload(app_client, token)
    assert [
        row["values"]["factory"] for row in resolved_payload["rows"]
    ] == ["Merged Supplier", "Merged Supplier"]
    confirm_response = app_client.post("/upload/confirm", data={"token": token})
    assert confirm_response.status_code == 200

    with sqlite3.connect(tmp_path / "gts-test.sqlite3") as connection:
        supplier_count = connection.execute("SELECT COUNT(*) FROM suppliers").fetchone()[0]
        quotation_rows = connection.execute(
            "SELECT factory, supplier_id FROM quotation_items ORDER BY gts_no"
        ).fetchall()
        aliases_text = connection.execute(
            "SELECT aliases_text FROM suppliers WHERE supplier_short_name = 'Merged Supplier'"
        ).fetchone()[0]
    assert supplier_count == 1
    assert quotation_rows[0][1] == quotation_rows[1][1]
    assert quotation_rows[0][0] == "Merged Supplier"
    assert quotation_rows[1][0] == "Merged Supplier"
    assert aliases_text == ""


def test_blank_factory_group_requires_resolution_and_import_fills_factory(
    app_client: TestClient,
    tmp_path: Path,
) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "blank-factory.xlsx",
                build_quotation_workbook(
                    [
                        {"gts_no": "GTS-BLANK-001", "factory": "", "unit": "pc", "unit_price": 10},
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    token = extract_token(upload_response.text)
    app_client.get(f"/upload/preview/stream/{token}")
    payload = preview_payload(app_client, token)
    blank_match = payload["supplier_matches"][0]

    assert blank_match["status"] == "blank_pending"
    assert blank_match["display_factory"] == "空白供应商"

    create_missing_name_response = app_client.post(
        "/upload/preview/supplier/batch",
        data={
            "token": token,
            "operator_name": "Nancy",
            f"action__{blank_match['key']}": "create",
        },
    )
    assert create_missing_name_response.status_code == 400
    assert "请填写供应商简称" in create_missing_name_response.text

    create_response = app_client.post(
        "/upload/preview/supplier/batch",
        data={
            "token": token,
            "operator_name": "Nancy",
            f"action__{blank_match['key']}": "create",
            f"supplier_short_name__{blank_match['key']}": "Blank Factory Short",
        },
        follow_redirects=False,
    )
    assert create_response.status_code == 303
    confirm_response = app_client.post("/upload/confirm", data={"token": token})
    assert confirm_response.status_code == 200
    with sqlite3.connect(tmp_path / "gts-test.sqlite3") as connection:
        row = connection.execute(
            "SELECT factory, supplier_id FROM quotation_items WHERE gts_no_normalized = 'GTSBLANK001'"
        ).fetchone()
    assert row[0] == "Blank Factory Short"
    assert row[1] is not None


def test_supplier_full_short_and_alias_auto_match_in_upload_preview(
    app_client: TestClient,
) -> None:
    create_supplier_for_test(
        app_client,
        full_name="Auto Full Name",
        short_name="Auto Short",
        aliases_text="Auto Alias",
    )
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "auto-match-suppliers.xlsx",
                build_quotation_workbook(
                    [
                        {"gts_no": "GTS-FULL", "factory": "Auto Full Name", "unit": "pc", "unit_price": 10},
                        {"gts_no": "GTS-SHORT", "factory": "Auto Short", "unit": "pc", "unit_price": 11},
                        {"gts_no": "GTS-ALIAS", "factory": "Auto Alias", "unit": "pc", "unit_price": 12},
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    token = extract_token(upload_response.text)
    app_client.get(f"/upload/preview/stream/{token}")
    payload = preview_payload(app_client, token)

    assert [match["status"] for match in payload["supplier_matches"]] == [
        "auto_matched",
        "auto_matched",
        "auto_matched",
    ]
    preview_page = app_client.get(f"/upload/preview/{token}")
    assert preview_page.status_code == 200
    assert "supplier-status-auto_matched" not in preview_page.text
    assert "所有供应商已匹配，可以导入。" in preview_page.text
    assert app_client.post("/upload/confirm", data={"token": token}).status_code == 200


def test_ambiguous_supplier_blocks_until_selected(app_client: TestClient) -> None:
    create_supplier_for_test(app_client, full_name="Ambiguous One", short_name="Ambiguous One", aliases_text="Shared Supplier")
    create_supplier_for_test(app_client, full_name="Ambiguous Two", short_name="Ambiguous Two", aliases_text="Shared Supplier")
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "ambiguous-supplier.xlsx",
                build_quotation_workbook(
                    [{"gts_no": "GTS-AMB-001", "factory": "Shared Supplier", "unit": "pc", "unit_price": 10}]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    token = extract_token(upload_response.text)
    app_client.get(f"/upload/preview/stream/{token}")
    payload = preview_payload(app_client, token)
    match = payload["supplier_matches"][0]

    assert match["status"] == "ambiguous_pending"
    assert len(match["candidate_suppliers"]) == 2
    assert app_client.post("/upload/confirm", data={"token": token}).status_code == 400

    response = app_client.post(
        "/upload/preview/supplier/batch",
        data={
            "token": token,
            "operator_name": "Nancy",
            f"action__{match['key']}": "ambiguous",
            f"supplier_id__{match['key']}": match["candidate_suppliers"][0]["id"],
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert app_client.post("/upload/confirm", data={"token": token}).status_code == 200


def test_upload_confirm_stops_before_import_when_auto_backup_fails(
    app_client: TestClient,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from app.routes import upload as upload_route
    from app.services.backup import BackupError

    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "backup-failure.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "gts_no": "GTS-BACKUP-FAIL",
                            "oem": "OEM-BACKUP-FAIL",
                            "factory": "Factory A",
                            "unit": "pc",
                            "unit_price": 10,
                        }
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 200
    token = extract_token(upload_response.text)
    stream_response = app_client.get(f"/upload/preview/stream/{token}")
    assert '"has_errors": false' in stream_response.text

    def fail_backup(reason: str):
        raise BackupError("自动备份失败，请检查 backups/auto 文件夹权限。")

    resolve_all_unmatched_suppliers(app_client, token)
    monkeypatch.setattr(upload_route, "create_auto_backup", fail_backup)
    confirm_response = app_client.post("/upload/confirm", data={"token": token})

    assert confirm_response.status_code == 500
    assert "自动备份失败" in confirm_response.text
    with sqlite3.connect(tmp_path / "gts-test.sqlite3") as connection:
        product_count = connection.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        upload_log_count = connection.execute(
            "SELECT COUNT(*) FROM operation_logs WHERE action_type = 'upload_full_quotation'"
        ).fetchone()[0]
    assert product_count == 0
    assert upload_log_count == 0


def test_generate_preview_highlights_multiple_candidates_and_shows_comments(
    app_client: TestClient,
) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "multiple-candidates.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "no": 1,
                            "gts_no": "GTS-MULTI-001",
                            "oem": "OEM-MULTI",
                            "factory": "Factory A",
                            "quantity": 1,
                            "unit": "pc",
                            "unit_price": 100,
                            "comment": "old stock",
                        },
                        {
                            "no": 2,
                            "gts_no": "GTS-MULTI-001",
                            "oem": "OEM-MULTI",
                            "factory": "Factory B",
                            "quantity": 1,
                            "unit": "pc",
                            "unit_price": 120,
                            "comment": "new stock",
                        },
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    upload_token = extract_token(upload_response.text)
    stream_response = app_client.get(f"/upload/preview/stream/{upload_token}")
    assert '"has_errors": false' in stream_response.text
    confirm_response = confirm_upload_after_creating_suppliers(app_client, upload_token)
    assert confirm_response.status_code == 200

    generate_response = app_client.post(
        "/generate/preview",
        data={"operator_name": "Nancy"},
        files={
            "request_file": (
                "request.xlsx",
                build_request_workbook([{"gts_no": "GTSMULTI001"}]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert generate_response.status_code == 200
    assert 'class="multiple-candidates-row"' in generate_response.text
    assert "candidate-line-review" in generate_response.text
    assert "Factory A" in generate_response.text
    assert "Factory B" in generate_response.text
    assert "old stock" in generate_response.text
    assert "new stock" in generate_response.text


def test_generate_download_allows_missing_quantity_and_leaves_total_blank(
    app_client: TestClient,
    tmp_path: Path,
) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "quotation-no-quantity-request.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "no": 1,
                            "gts_no": "GTS-NO-QTY-001",
                            "oem": "OEM-NO-QTY",
                            "factory": "Factory A",
                            "quantity": 1,
                            "unit": "pc",
                            "unit_price": 88,
                            "total_price": 88,
                        }
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    upload_token = extract_token(upload_response.text)
    stream_response = app_client.get(f"/upload/preview/stream/{upload_token}")
    assert '"has_errors": false' in stream_response.text
    confirm_response = confirm_upload_after_creating_suppliers(app_client, upload_token)
    assert confirm_response.status_code == 200

    generate_response = app_client.post(
        "/generate/preview",
        data={"operator_name": "Nancy"},
        files={
            "request_file": (
                "request-no-quantity.xlsx",
                build_request_workbook([{"gts_no": "GTSNOQTY001"}]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert generate_response.status_code == 200
    assert "未填写数量" in generate_response.text

    generate_token = extract_token(generate_response.text)
    candidate_id = fetch_single_candidate_id(tmp_path / "gts-test.sqlite3")
    download_response = app_client.post(
        "/generate/download",
        data={
            "token": generate_token,
            "include__2": "1",
            "candidate__2": str(candidate_id),
        },
    )

    assert download_response.status_code == 200
    workbook = load_workbook(BytesIO(download_response.content))
    worksheet = workbook.active
    assert worksheet["B3"].value == "GTS-NO-QTY-001"
    assert worksheet["H3"].value is None
    assert worksheet["K3"].value is None


def test_generate_preview_refuses_when_all_rows_have_no_gts_or_oem(
    app_client: TestClient,
) -> None:
    generate_response = app_client.post(
        "/generate/preview",
        data={"operator_name": "Nancy"},
        files={
            "request_file": (
                "request-no-identifiers.xlsx",
                build_request_workbook(
                    [
                        {"description": "Only description"},
                        {"quantity": 5, "comment": "No GTS or OEM"},
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert generate_response.status_code == 400
    assert "需求文件中没有可识别的 GTS 或 OEM" in generate_response.text


def test_generate_preview_allows_mixed_valid_and_invalid_request_rows(
    app_client: TestClient,
    tmp_path: Path,
) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "quotation-mixed-request.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "no": 1,
                            "gts_no": "GTS-MIXED-001",
                            "oem": "OEM-MIXED",
                            "factory": "Factory A",
                            "quantity": 1,
                            "unit": "pc",
                            "unit_price": 66,
                        }
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    upload_token = extract_token(upload_response.text)
    stream_response = app_client.get(f"/upload/preview/stream/{upload_token}")
    assert '"has_errors": false' in stream_response.text
    confirm_response = confirm_upload_after_creating_suppliers(app_client, upload_token)
    assert confirm_response.status_code == 200

    generate_response = app_client.post(
        "/generate/preview",
        data={"operator_name": "Nancy"},
        files={
            "request_file": (
                "request-mixed.xlsx",
                build_request_workbook(
                    [
                        {"description": "Invalid request row", "quantity": 2},
                        {"gts_no": "GTSMIXED001", "quantity": 4},
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert generate_response.status_code == 200
    assert "GTS 和 OEM 都缺失" in generate_response.text
    assert 'class="missing-identifier-row"' in generate_response.text
    assert "GTS-MIXED-001" in generate_response.text

    generate_token = extract_token(generate_response.text)
    candidate_id = fetch_single_candidate_id(tmp_path / "gts-test.sqlite3")
    download_response = app_client.post(
        "/generate/download",
        data={
            "token": generate_token,
            "include__2": "1",
            "candidate__2": "-1",
            "include__3": "1",
            "candidate__3": str(candidate_id),
        },
    )

    assert download_response.status_code == 200
    workbook = load_workbook(BytesIO(download_response.content))
    worksheet = workbook.active
    assert worksheet["C3"].value == "Invalid request row"
    assert worksheet["H3"].value == 2
    assert worksheet["B4"].value == "GTS-MIXED-001"
    assert worksheet["H4"].value == 4
    assert worksheet.max_row == 4


def test_hs_code_upload_overwrites_search_displays_and_export_keeps_order(
    app_client: TestClient,
) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "hs-products.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "gts_no": "GTS-HS-001",
                            "oem": "OEM-HS-001",
                            "factory": "Factory A",
                            "unit": "pc",
                            "unit_price": 10,
                        },
                        {
                            "gts_no": "GTS-HS-002",
                            "oem": "OEM-HS-002",
                            "factory": "Factory B",
                            "unit": "pc",
                            "unit_price": 20,
                        },
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    upload_token = extract_token(upload_response.text)
    app_client.get(f"/upload/preview/stream/{upload_token}")
    confirm_response = confirm_upload_after_creating_suppliers(app_client, upload_token)
    assert confirm_response.status_code == 200

    hs_upload_response = app_client.post(
        "/hs-codes/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "hs-upload.xlsx",
                build_hs_code_workbook(
                    [
                        {"gts_no": "GTS-HS-001", "hs_code": "87089910"},
                        {"gts_no": "GTS-HS-002", "hs_code": "87089920"},
                    ],
                    hs_header="海关编码",
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert hs_upload_response.status_code == 200
    assert "87089910" in hs_upload_response.text
    hs_token = extract_token(hs_upload_response.text)
    hs_confirm_response = app_client.post("/hs-codes/upload/confirm", data={"token": hs_token})
    assert hs_confirm_response.status_code == 200
    assert "已更新" in hs_confirm_response.text
    assert not (app_client.upload_path / f"hs_upload_preview_{hs_token}.json").exists()

    overwrite_response = app_client.post(
        "/hs-codes/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "hs-overwrite.xlsx",
                build_hs_code_workbook(
                    [{"gts_no": "GTS-HS-001", "hs_code": "87089999"}],
                    hs_header="HS",
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    overwrite_token = extract_token(overwrite_response.text)
    overwrite_confirm = app_client.post(
        "/hs-codes/upload/confirm",
        data={"token": overwrite_token},
    )
    assert overwrite_confirm.status_code == 200

    search_response = app_client.get("/search", params={"field": "gts_no", "q": "GTSHS001"})
    assert search_response.status_code == 200
    assert "87089999" in search_response.text
    assert "87089910" not in search_response.text

    generate_response = app_client.post(
        "/hs-codes/generate/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "hs-request.xlsx",
                build_hs_request_workbook(
                    [
                        {"gts_no": "GTS-HS-002"},
                        {"gts_no": "GTS-HS-001"},
                        {"gts_no": "GTS-HS-MISSING"},
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert generate_response.status_code == 200
    generate_token = extract_token(generate_response.text)
    download_response = app_client.post(
        "/hs-codes/generate/download",
        data={"token": generate_token},
    )
    assert download_response.status_code == 200
    assert download_response.headers["content-disposition"].startswith(
        'attachment; filename="Nancy-hs-'
    )
    assert not (
        app_client.upload_path / f"hs_generate_preview_{generate_token}.json"
    ).exists()

    workbook = load_workbook(BytesIO(download_response.content))
    worksheet = workbook.active
    assert [worksheet.cell(row=1, column=index).value for index in range(1, 4)] == [
        "GTS",
        "OEM",
        "HS Code",
    ]
    assert [worksheet.cell(row=row, column=1).value for row in range(2, 5)] == [
        "GTS-HS-002",
        "GTS-HS-001",
        "GTS-HS-MISSING",
    ]
    assert worksheet["B2"].value == "OEM-HS-002"
    assert worksheet["C2"].value == "87089920"
    assert worksheet["B3"].value == "OEM-HS-001"
    assert worksheet["C3"].value == "87089999"
    assert worksheet["B4"].value is None
    assert worksheet["C4"].value is None

    logs_response = app_client.get("/logs")
    assert logs_response.status_code == 200
    assert "update_hs_code" in logs_response.text
    assert "generate_hs_code" in logs_response.text


def test_product_edit_updates_current_fields_used_by_quotation_and_hs_exports(
    app_client: TestClient,
    tmp_path: Path,
) -> None:
    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "product-edit-source.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "gts_no": "GTS-OLD-001",
                            "description": "Old product description",
                            "chinese_description": "旧品名",
                            "oem": "OEM-OLD-001",
                            "factory": "Factory A",
                            "unit": "pc",
                            "unit_price": 99,
                        }
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    upload_token = extract_token(upload_response.text)
    app_client.get(f"/upload/preview/stream/{upload_token}")
    confirm_response = confirm_upload_after_creating_suppliers(app_client, upload_token)
    assert confirm_response.status_code == 200

    edit_page = app_client.get("/products/1/edit")
    assert edit_page.status_code == 200
    assert "data-product-edit-form" in edit_page.text
    assert 'value="GTS-OLD-001"' in edit_page.text
    assert 'data-original-value="GTS-OLD-001"' in edit_page.text
    assert 'value="OEM-OLD-001"' in edit_page.text
    assert 'value="Old product description"' in edit_page.text
    assert '<label for="chinese_description">品名</label>' in edit_page.text
    assert 'value="旧品名"' in edit_page.text
    assert "data-product-edit-submit disabled" in edit_page.text

    bad_password_response = app_client.post(
        "/products/1/edit",
        data={
            "operator_name": "Nancy",
            "gts_no": "GTS-NEW-001",
            "oem": "OEM-NEW-001",
            "description": "New product description",
            "chinese_description": "新品名",
            "hs_code": "87089999",
            "edit_password": "wrong",
        },
    )
    assert bad_password_response.status_code == 400
    assert "确认密码不正确" in bad_password_response.text
    assert 'value="GTS-NEW-001"' in bad_password_response.text
    assert 'value="新品名"' in bad_password_response.text
    assert 'data-original-value="GTS-OLD-001"' in bad_password_response.text

    edit_response = app_client.post(
        "/products/1/edit",
        data={
            "operator_name": "Nancy",
            "gts_no": "GTS-NEW-001",
            "oem": "OEM-NEW-001",
            "description": "New product description",
            "chinese_description": "新品名",
            "hs_code": "87089999",
            "edit_password": "55123511",
        },
    )
    assert edit_response.status_code == 200
    assert "产品资料已更新" in edit_response.text

    generate_response = app_client.post(
        "/generate/preview",
        data={"operator_name": "Nancy"},
        files={
            "request_file": (
                "old-request.xlsx",
                build_request_workbook(
                    [
                        {
                            "gts_no": "GTS-OLD-001",
                            "description": "Request description should not override product",
                            "oem": "OEM-OLD-001",
                            "quantity": 2,
                        }
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert generate_response.status_code == 200
    assert "GTS 已从 GTS-OLD-001 改为 GTS-NEW-001" in generate_response.text
    assert "OEM 已从 OEM-OLD-001 改为 OEM-NEW-001" in generate_response.text

    generate_token = extract_token(generate_response.text)
    candidate_id = fetch_single_candidate_id(tmp_path / "gts-test.sqlite3")
    download_response = app_client.post(
        "/generate/download",
        data={
            "token": generate_token,
            "include__2": "1",
            "candidate__2": str(candidate_id),
        },
    )
    assert download_response.status_code == 200
    workbook = load_workbook(BytesIO(download_response.content))
    worksheet = workbook.active
    assert worksheet["B3"].value == "GTS-NEW-001"
    assert worksheet["C3"].value == "New product description"
    assert worksheet["D3"].value == "OEM-NEW-001"
    assert worksheet["G3"].value == "新品名"

    hs_generate_response = app_client.post(
        "/hs-codes/generate/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "old-hs-request.xlsx",
                build_hs_request_workbook([{"gts_no": "GTS-OLD-001"}]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert hs_generate_response.status_code == 200
    assert "GTS 已从 GTS-OLD-001 改为 GTS-NEW-001" in hs_generate_response.text

    hs_token = extract_token(hs_generate_response.text)
    hs_download_response = app_client.post(
        "/hs-codes/generate/download",
        data={"token": hs_token},
    )
    assert hs_download_response.status_code == 200
    hs_workbook = load_workbook(BytesIO(hs_download_response.content))
    hs_worksheet = hs_workbook.active
    assert hs_worksheet["A2"].value == "GTS-NEW-001"
    assert hs_worksheet["B2"].value == "OEM-NEW-001"
    assert hs_worksheet["C2"].value == "87089999"

    logs_response = app_client.get("/logs")
    assert logs_response.status_code == 200
    assert "编辑产品" in logs_response.text


def test_operator_name_can_be_saved_changed_and_prefilled_in_session(
    app_client: TestClient,
) -> None:
    operator_response = app_client.post("/operator", data={"operator_name": "Alice"})
    assert operator_response.status_code == 200
    assert "操作人：Alice" in operator_response.text
    assert "修改操作人" in operator_response.text
    assert 'data-operator-modal' in operator_response.text
    assert 'action="/logout"' in operator_response.text

    upload_page = app_client.get("/upload")
    assert upload_page.status_code == 200
    assert 'name="operator_name" required autocomplete="name" value="Alice"' in upload_page.text

    generate_page = app_client.get("/generate")
    assert generate_page.status_code == 200
    assert 'name="operator_name" required autocomplete="name" value="Alice"' in generate_page.text

    hs_upload_page = app_client.get("/hs-codes/upload")
    assert hs_upload_page.status_code == 200
    assert 'name="operator_name" required autocomplete="name" value="Alice"' in hs_upload_page.text

    hs_generate_page = app_client.get("/hs-codes/generate")
    assert hs_generate_page.status_code == 200
    assert 'name="operator_name" required autocomplete="name" value="Alice"' in hs_generate_page.text

    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Bob"},
        files={
            "excel_file": (
                "operator-session.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "gts_no": "GTS-OP-001",
                            "oem": "OEM-OP-001",
                            "factory": "Factory A",
                            "unit": "pc",
                            "unit_price": 12,
                        }
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 200
    upload_token = extract_token(upload_response.text)
    app_client.get(f"/upload/preview/stream/{upload_token}")
    confirm_response = confirm_upload_after_creating_suppliers(app_client, upload_token)
    assert confirm_response.status_code == 200

    updated_upload_page = app_client.get("/upload")
    assert updated_upload_page.status_code == 200
    assert "操作人：Bob" in updated_upload_page.text
    assert 'name="operator_name" required autocomplete="name" value="Bob"' in updated_upload_page.text

    edit_page = app_client.get("/products/1/edit")
    assert edit_page.status_code == 200
    assert 'name="operator_name" required autocomplete="off" value="Bob"' in edit_page.text


def test_supplier_create_edit_search_and_import_linking(
    app_client: TestClient,
    tmp_path: Path,
) -> None:
    create_response = app_client.post(
        "/suppliers/new",
        data={
            "operator_name": "Nancy",
            "supplier_full_name": "Factory A Full",
            "supplier_short_name": "Factory A",
            "contact_person": "Alice",
            "phone": "123",
            "wechat": "alice-wx",
            "city": "Ningbo",
            "province": "Zhejiang",
            "product_scope": "Mirror",
            "factory_or_trader": "工厂",
            "quality_level": "A",
            "price_level": "中",
            "notes": "Initial",
            "edit_password": "55123511",
        },
    )
    assert create_response.status_code == 200
    assert "Factory A" in create_response.text
    assert "Alice" in create_response.text
    assert 'class="supplier-detail-form"' not in create_response.text
    assert 'href="/suppliers/1/edit?mode=edit"' in create_response.text

    duplicate_short_name_response = app_client.post(
        "/suppliers/new",
        data={
            "operator_name": "Nancy",
            "supplier_full_name": "Another Factory",
            "supplier_short_name": "Factory A",
            "edit_password": "55123511",
        },
    )
    assert duplicate_short_name_response.status_code == 400
    assert "供应商简称已存在" in duplicate_short_name_response.text

    search_supplier_response = app_client.get("/suppliers", params={"q": "Mirror"})
    assert search_supplier_response.status_code == 200
    assert "Factory A" in search_supplier_response.text
    assert "Ningbo" in search_supplier_response.text
    assert "更多" in search_supplier_response.text
    assert "<th>联系人</th>" not in search_supplier_response.text
    assert "<th>质量评分</th>" not in search_supplier_response.text
    assert "缺供应商信息" in search_supplier_response.text
    assert "缺评分" in search_supplier_response.text
    assert "data-status-tag-supplier-info" in search_supplier_response.text
    assert "data-status-tag-rating" in search_supplier_response.text

    edit_response = app_client.post(
        "/suppliers/1/edit",
        data={
            "operator_name": "Nancy",
            "supplier_full_name": "Factory A Full Updated",
            "supplier_short_name": "Factory A Updated",
            "contact_person": "Alice",
            "phone": "456",
            "wechat": "alice-wx",
            "city": "Ningbo",
            "province": "Zhejiang",
            "product_scope": "Mirror, Lamp",
            "factory_or_trader": "工厂",
            "quality_level": "A",
            "price_level": "中",
            "notes": "Updated",
            "edit_password": "55123511",
        },
    )
    assert edit_response.status_code == 200
    assert "Factory A Updated" in edit_response.text
    assert 'class="supplier-detail-form"' not in edit_response.text

    edit_form = app_client.get("/suppliers/1/edit", params={"mode": "edit"})
    assert edit_form.status_code == 200
    assert 'class="supplier-detail-form"' in edit_form.text
    assert "首页" in edit_form.text
    assert "供应商" in edit_form.text
    assert "Factory A Updated" in edit_form.text

    upload_response = app_client.post(
        "/upload/preview",
        data={"operator_name": "Nancy"},
        files={
            "excel_file": (
                "supplier-link.xlsx",
                build_quotation_workbook(
                    [
                        {
                            "gts_no": "GTS-SUP-001",
                            "oem": "OEM-SUP-001",
                            "factory": "Factory A Updated",
                            "unit": "pc",
                            "unit_price": 88,
                        },
                        {
                            "gts_no": "GTS-SUP-002",
                            "oem": "OEM-SUP-002",
                            "factory": "Unlinked Factory",
                            "unit": "pc",
                            "unit_price": 99,
                        },
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 200
    token = extract_token(upload_response.text)
    app_client.get(f"/upload/preview/stream/{token}")
    resolve_all_unmatched_suppliers(app_client, token)
    confirm_response = app_client.post("/upload/confirm", data={"token": token})
    assert confirm_response.status_code == 200

    with sqlite3.connect(tmp_path / "gts-test.sqlite3") as connection:
        connection.row_factory = sqlite3.Row
        linked = connection.execute(
            """
            SELECT q.supplier_id, s.supplier_short_name
            FROM quotation_items q
            LEFT JOIN suppliers s ON s.id = q.supplier_id
            WHERE q.gts_no_normalized = 'GTSSUP001'
            """
        ).fetchone()
        fallback = connection.execute(
            """
            SELECT q.supplier_id, q.factory, s.supplier_short_name
            FROM quotation_items q
            LEFT JOIN suppliers s ON s.id = q.supplier_id
            WHERE gts_no_normalized = 'GTSSUP002'
            """
        ).fetchone()
    assert linked["supplier_id"] == 1
    assert linked["supplier_short_name"] == "Factory A Updated"
    assert fallback["supplier_id"] is not None
    assert fallback["factory"] == "Unlinked Factory"
    assert fallback["supplier_short_name"] == "Unlinked Factory"

    linked_search = app_client.get("/search", params={"field": "gts_no", "q": "GTSSUP001"})
    fallback_search = app_client.get("/search", params={"field": "gts_no", "q": "GTSSUP002"})
    assert "Factory A Updated" in linked_search.text
    assert "Unlinked Factory" in fallback_search.text


def test_login_rejects_wrong_shared_access_code(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("SHARED_ACCESS_CODE", ACCESS_CODE)
    monkeypatch.setenv("SESSION_SECRET_KEY", "test-session-secret-key")
    monkeypatch.setenv("PRODUCT_EDIT_PASSWORD", "55123511")
    monkeypatch.setenv("DATABASE_PATH", str(tmp_path / "auth-test.sqlite3"))

    from app.config import get_settings
    from app.main import create_app

    get_settings.cache_clear()
    client = TestClient(create_app())

    protected_response = client.get("/upload", follow_redirects=False)
    assert protected_response.status_code == 303
    assert protected_response.headers["location"] == "/login"

    login_page = client.get("/login")
    assert login_page.status_code == 200
    assert "操作人" in login_page.text
    assert 'name="operator_name"' in login_page.text

    bad_login_response = client.post("/login", data={"access_code": "wrong-code"})
    assert bad_login_response.status_code == 401
    assert "访问码不正确" in bad_login_response.text

    good_login_response = client.post(
        "/login",
        data={"access_code": ACCESS_CODE, "operator_name": "Nancy"},
        follow_redirects=False,
    )
    assert good_login_response.status_code == 303

    home_response = client.get("/")
    assert home_response.status_code == 200
    assert "操作人：Nancy" in home_response.text


def build_quotation_workbook(rows: list[dict]) -> BytesIO:
    headers = [
        "No.",
        "GTS No.",
        "Description",
        "OEM",
        "Photo",
        "Factory",
        "Chinese Description",
        "Quantity",
        "Unit",
        "Unit Price",
        "Total Price",
        "Item/Package",
        "Packages",
        "Weight / Package",
        "G.W.",
        "Length",
        "Width",
        "Height",
        "Measurements / Volume",
        "Packaging",
        "Expected Delivery",
        "Comment",
    ]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Quotation"
    worksheet["A1"] = "GTS Internal Tool Upload Template"
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=3, column=column_index, value=header)

    field_order = [
        "no",
        "gts_no",
        "description",
        "oem",
        "photo",
        "factory",
        "chinese_description",
        "quantity",
        "unit",
        "unit_price",
        "total_price",
        "item_per_package",
        "packages",
        "weight_per_package",
        "gross_weight",
        "length",
        "width",
        "height",
        "measurements_volume",
        "packaging",
        "expected_delivery",
        "comment",
    ]
    for row_index, row in enumerate(rows, start=4):
        for column_index, field in enumerate(field_order, start=1):
            worksheet.cell(row=row_index, column=column_index, value=row.get(field, ""))
    return workbook_bytes(workbook)


def build_request_workbook(rows: list[dict]) -> BytesIO:
    headers = ["GTS No.", "Description", "OEM", "Quantity", "Unit", "Comment"]
    field_order = ["gts_no", "description", "oem", "quantity", "unit", "comment"]
    workbook = Workbook()
    worksheet = workbook.active
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column_index, value=header)
    for row_index, row in enumerate(rows, start=2):
        for column_index, field in enumerate(field_order, start=1):
            worksheet.cell(row=row_index, column=column_index, value=row.get(field, ""))
    return workbook_bytes(workbook)


def build_hs_code_workbook(rows: list[dict], hs_header: str = "HS Code") -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    headers = ["GTS No.", "Description", hs_header, "Comment"]
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column_index, value=header)
    for row_index, row in enumerate(rows, start=2):
        worksheet.cell(row=row_index, column=1, value=row.get("gts_no", ""))
        worksheet.cell(row=row_index, column=2, value=row.get("description", ""))
        worksheet.cell(row=row_index, column=3, value=row.get("hs_code", ""))
        worksheet.cell(row=row_index, column=4, value=row.get("comment", ""))
    return workbook_bytes(workbook)


def build_hs_request_workbook(rows: list[dict]) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=1, value="GTS")
    for row_index, row in enumerate(rows, start=2):
        worksheet.cell(row=row_index, column=1, value=row.get("gts_no", ""))
    return workbook_bytes(workbook)


def workbook_bytes(workbook: Workbook) -> BytesIO:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def extract_token(html: str) -> str:
    match = re.search(r'name="token" value="([^"]+)"', html)
    assert match is not None
    return match.group(1)


def preview_payload(client: TestClient, token: str) -> dict:
    return json.loads((client.upload_path / f"preview_{token}.json").read_text())


def confirm_upload_after_creating_suppliers(client: TestClient, token: str):
    resolve_all_unmatched_suppliers(client, token)
    return client.post("/upload/confirm", data={"token": token})


def resolve_all_unmatched_suppliers(client: TestClient, token: str) -> None:
    payload = preview_payload(client, token)
    data = {"token": token, "operator_name": payload["operator_name"]}
    for match in payload.get("supplier_matches") or []:
        if match["status"] not in {"pending_unmatched", "blank_pending", "ambiguous_pending"}:
            continue
        factory = match.get("factory") or f"Blank Supplier {match['key']}"
        data[f"action__{match['key']}"] = "create"
        data[f"supplier_short_name__{match['key']}"] = factory
    if len(data) == 2:
        return
    response = client.post(
        "/upload/preview/supplier/batch",
        data=data,
        follow_redirects=False,
    )
    assert response.status_code == 303


def create_supplier_for_test(
    client: TestClient,
    *,
    full_name: str,
    short_name: str,
    aliases_text: str = "",
) -> None:
    response = client.post(
        "/suppliers/new",
        data={
            "operator_name": "Nancy",
            "supplier_full_name": full_name,
            "supplier_short_name": short_name,
            "aliases_text": aliases_text,
            "edit_password": "55123511",
        },
    )
    assert response.status_code == 200


def fetch_single_candidate_id(database_path: Path) -> int:
    with sqlite3.connect(database_path) as connection:
        row = connection.execute("SELECT id FROM quotation_items").fetchone()
        assert row is not None
        return int(row[0])


def assert_breadcrumb(html: str, labels: list[str]) -> None:
    match = re.search(
        r'<ol class="breadcrumb-list">(.*?)</ol>',
        html,
        flags=re.DOTALL,
    )
    assert match is not None
    breadcrumb_text = re.sub(r"<[^>]+>", " ", match.group(1))
    breadcrumb_text = re.sub(r"\s+", " ", breadcrumb_text).strip()
    assert " ".join(labels) == breadcrumb_text
