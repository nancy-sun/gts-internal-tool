import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from app.config import BASE_DIR
from app.services.excel_parser import (
    HEADER_ALIASES,
    _clean_cell_value,
    _parse_number,
    normalize_header_label,
)
from app.services.normalization import normalize_gts_no, normalize_oem


REQUEST_FIELDS = ["gts_no", "description", "oem", "quantity", "comment"]
REQUEST_HEADER_ALIASES = {
    field: HEADER_ALIASES[field]
    for field in ("gts_no", "description", "oem", "quantity", "comment")
}
REQUEST_HEADER_LOOKUP = {
    normalize_header_label(alias): field
    for field, aliases in REQUEST_HEADER_ALIASES.items()
    for alias in aliases
}


@dataclass
class ParsedRequestRow:
    row_number: int
    values: dict[str, Any]
    warnings: list[str]
    errors: list[str]

    @property
    def is_valid(self) -> bool:
        return not self.errors


def load_request_template_config(config_path: Path | None = None) -> dict[str, Any]:
    path = config_path or BASE_DIR / "config" / "request_template.json"
    with path.open("r", encoding="utf-8") as config_file:
        return json.load(config_file)


def parse_request_workbook(
    workbook_path: Path,
    config: dict[str, Any] | None = None,
) -> list[ParsedRequestRow]:
    template = config or load_request_template_config()
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet_name = template.get("sheet_name")
        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        header_row = resolve_request_header_row(worksheet, template)
        start_row = header_row + 1
        max_rows = int(template.get("max_rows", 300))
        columns = resolve_request_columns(worksheet, header_row, template)

        parsed_rows: list[ParsedRequestRow] = []
        for row_number in range(start_row, start_row + max_rows):
            values = {
                field: _clean_cell_value(worksheet[f"{columns[field]}{row_number}"].value)
                if field in columns
                else ""
                for field in REQUEST_FIELDS
            }
            if all(value in ("", None) for value in values.values()):
                continue

            warnings: list[str] = []
            errors: list[str] = []
            gts_no_normalized, gts_warnings = normalize_gts_no(values["gts_no"])
            oem_normalized, oem_warnings = normalize_oem(values["oem"])
            values["gts_no_normalized"] = gts_no_normalized
            values["oem_normalized"] = oem_normalized
            values["quantity"] = _parse_number(values["quantity"], "quantity", warnings)
            warnings.extend([f"GTS No. {warning}" for warning in gts_warnings])
            warnings.extend([f"OEM {warning}" for warning in oem_warnings])

            if not gts_no_normalized and not oem_normalized:
                errors.append("Row must contain GTS No. or OEM.")

            parsed_rows.append(
                ParsedRequestRow(
                    row_number=row_number,
                    values=values,
                    warnings=warnings,
                    errors=errors,
                )
            )
        return parsed_rows
    finally:
        workbook.close()


def resolve_request_header_row(worksheet, template: dict[str, Any]) -> int:
    max_header_scan_rows = int(template.get("header_scan_rows", 100))
    max_header_scan_columns = int(template.get("header_scan_columns", 40))
    max_row = worksheet.max_row or max_header_scan_rows
    scan_row_limit = min(max_row, max_header_scan_rows)

    for row in range(1, scan_row_limit + 1):
        for column in range(1, max_header_scan_columns + 1):
            header = normalize_header_label(worksheet.cell(row=row, column=column).value)
            if header in REQUEST_HEADER_LOOKUP:
                return row
    return int(template.get("header_row", 1))


def resolve_request_columns(worksheet, header_row: int, template: dict[str, Any]) -> dict[str, str]:
    max_header_scan_columns = int(template.get("header_scan_columns", 40))
    detected_headers: dict[str, str] = {}
    for column_index in range(1, max_header_scan_columns + 1):
        normalized_header = normalize_header_label(
            worksheet.cell(row=header_row, column=column_index).value
        )
        field = REQUEST_HEADER_LOOKUP.get(normalized_header)
        if field:
            detected_headers.setdefault(field, get_column_letter(column_index))

    fallback_columns = template.get("columns", {})
    return {
        field: detected_headers[field]
        for field in REQUEST_FIELDS
        if field in detected_headers
    } or {
        field: fallback_columns[field]
        for field in REQUEST_FIELDS
        if field in fallback_columns
    }
