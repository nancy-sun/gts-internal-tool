from __future__ import annotations

from dataclasses import asdict, dataclass
from io import BytesIO
from sqlite3 import Connection, Row
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.lookup_helpers import (
    fetch_products_by_normalized_field,
    unique_text_values,
)
from app.services.operation_logging import create_operation_log
from app.services.request_parser import ParsedRequestRow


GENERATED_COLUMNS = [
    ("no", "No."),
    ("gts_no", "GTS No."),
    ("description", "Description"),
    ("oem", "OEM"),
    ("photo", "Photo"),
    ("factory", "Factory"),
    ("chinese_description", "Chinese Description"),
    ("quantity", "Quantity"),
    ("unit", "Unit"),
    ("unit_price", "Unit Price"),
    ("total_price", "Total Price"),
    ("item_per_package", "Item/Package"),
    ("packages", "Packages"),
    ("weight_per_package", "Weight / Package"),
    ("gross_weight", "G.W."),
    ("length", "Length"),
    ("width", "Width"),
    ("height", "Height"),
    ("measurements_volume", "Measurements / Volume"),
    ("packaging", "Packaging"),
    ("expected_delivery", "Expected Delivery"),
    ("comment", "Comment"),
    ("updated_by", "Updated By"),
    ("updated_at", "Updated At"),
]
BLANK_ROW_CANDIDATE_ID = -1


@dataclass(frozen=True)
class GenerationLookupContext:
    products_by_gts: dict[str, Row]
    products_by_oem: dict[str, Row]
    products_by_historical_gts: dict[str, Row]
    products_by_historical_oem: dict[str, Row]
    candidates_by_product_id: dict[int, list[Row]]


def build_generation_preview(
    connection: Connection,
    parsed_rows: list[ParsedRequestRow],
) -> list[dict[str, Any]]:
    lookup_context = build_generation_lookup_context(connection, parsed_rows)
    preview_rows = []
    for parsed_row in parsed_rows:
        preview = asdict(parsed_row)
        preview["status"] = "ready"
        preview["product"] = None
        preview["candidates"] = []

        if parsed_row.errors:
            preview["status"] = "invalid"
            preview_rows.append(preview)
            continue
        if not parsed_row.values.get("gts_no_normalized") and not parsed_row.values.get("oem_normalized"):
            preview["status"] = "missing_identifier"
            preview_rows.append(preview)
            continue

        product, conflict = find_product_in_context(
            lookup_context,
            parsed_row.values,
        )
        if conflict:
            preview["status"] = "conflict"
            preview["errors"].append(
                "GTS 和 OEM 匹配到不同产品，需要人工确认。"
            )
            preview_rows.append(preview)
            continue

        if not product:
            preview["status"] = "unmatched"
            preview_rows.append(preview)
            continue

        preview["product"] = dict(product)
        candidates = lookup_context.candidates_by_product_id.get(product["id"], [])
        preview["candidates"] = [dict(candidate) for candidate in candidates]
        preview["change_notices"] = build_product_change_notices(
            product,
            parsed_row.values,
        )
        if not candidates:
            preview["status"] = "no_quotation"
        elif len(candidates) > 1:
            preview["status"] = "multiple_candidates"
        preview_rows.append(preview)
    return preview_rows


def build_generation_lookup_context(
    connection: Connection,
    parsed_rows: list[ParsedRequestRow],
) -> GenerationLookupContext:
    valid_values = [row.values for row in parsed_rows if not row.errors]
    gts_numbers = unique_text_values(row.get("gts_no_normalized") for row in valid_values)
    oem_numbers = unique_text_values(row.get("oem_normalized") for row in valid_values)

    products_by_gts = fetch_products_by_normalized_field(
        connection,
        "gts_no_normalized",
        gts_numbers,
        order_by="updated_at DESC, id DESC",
    )
    products_by_oem = fetch_products_by_normalized_field(
        connection,
        "oem_normalized",
        oem_numbers,
        order_by="updated_at DESC, id DESC",
    )
    products_by_historical_gts = fetch_products_by_historical_quotation_field(
        connection,
        "gts_no_normalized",
        gts_numbers,
    )
    products_by_historical_oem = fetch_products_by_historical_quotation_field(
        connection,
        "oem_normalized",
        oem_numbers,
    )
    product_ids = {
        int(product["id"])
        for product in [
            *products_by_gts.values(),
            *products_by_oem.values(),
            *products_by_historical_gts.values(),
            *products_by_historical_oem.values(),
        ]
    }
    candidates_by_product_id = fetch_quotation_candidates_by_product_id(
        connection,
        product_ids,
    )
    return GenerationLookupContext(
        products_by_gts=products_by_gts,
        products_by_oem=products_by_oem,
        products_by_historical_gts=products_by_historical_gts,
        products_by_historical_oem=products_by_historical_oem,
        candidates_by_product_id=candidates_by_product_id,
    )


def find_product_in_context(
    context: GenerationLookupContext,
    values: dict[str, Any],
) -> tuple[Row | None, bool]:
    gts_no_normalized = values.get("gts_no_normalized") or ""
    oem_normalized = values.get("oem_normalized") or ""
    gts_product = context.products_by_gts.get(
        gts_no_normalized,
    ) or context.products_by_historical_gts.get(gts_no_normalized)
    oem_product = context.products_by_oem.get(
        oem_normalized,
    ) or context.products_by_historical_oem.get(oem_normalized)

    if gts_product and oem_product and gts_product["id"] != oem_product["id"]:
        return None, True
    if gts_product:
        return gts_product, False
    if oem_product:
        return oem_product, False
    return None, False


def fetch_products_by_historical_quotation_field(
    connection: Connection,
    field: str,
    values: list[str],
) -> dict[str, Row]:
    if field not in {"gts_no_normalized", "oem_normalized"}:
        raise ValueError("Unsupported historical quotation lookup field")
    if not values:
        return {}

    placeholders = ", ".join("?" for _ in values)
    rows = connection.execute(
        f"""
        SELECT
            q.{field} AS lookup_value,
            p.*
        FROM quotation_items q
        JOIN products p ON p.id = q.product_id
        WHERE q.{field} IN ({placeholders})
        ORDER BY q.updated_at DESC, q.id DESC
        """,
        values,
    ).fetchall()
    products = {}
    for row in rows:
        lookup_value = _text(row["lookup_value"])
        if lookup_value and lookup_value not in products:
            products[lookup_value] = row
    return products


def build_product_change_notices(
    product: Row | dict[str, Any],
    request_values: dict[str, Any],
) -> list[str]:
    notices = []
    request_gts = _text(request_values.get("gts_no"))
    request_oem = _text(request_values.get("oem"))
    product_gts = _text(product["gts_no"])
    product_oem = _text(product["oem"])
    if request_gts and request_values.get("gts_no_normalized") != product["gts_no_normalized"]:
        notices.append(f"GTS 已从 {request_gts} 改为 {product_gts or '(空)'}")
    if request_oem and request_values.get("oem_normalized") != product["oem_normalized"]:
        notices.append(f"OEM 已从 {request_oem} 改为 {product_oem or '(空)'}")
    return notices


def fetch_quotation_candidates_by_product_id(
    connection: Connection,
    product_ids: set[int],
) -> dict[int, list[Row]]:
    if not product_ids:
        return {}

    ordered_ids = sorted(product_ids)
    placeholders = ", ".join("?" for _ in ordered_ids)
    rows = connection.execute(
        f"""
        SELECT *
        FROM quotation_items
        WHERE product_id IN ({placeholders})
        ORDER BY product_id, updated_at DESC, id DESC
        """,
        ordered_ids,
    ).fetchall()

    grouped_rows: dict[int, list[Row]] = {}
    for row in rows:
        grouped_rows.setdefault(int(row["product_id"]), []).append(row)

    return {
        product_id: dedupe_candidate_rows(product_rows)
        for product_id, product_rows in grouped_rows.items()
    }


def dedupe_candidate_rows(rows: list[Row]) -> list[Row]:
    candidates = []
    seen = set()
    for row in rows:
        signature = (
            _text(row["gts_no_normalized"]),
            _text(row["factory"]),
            _text(row["unit"]),
            _price_key(row["unit_price"]),
        )
        if signature in seen:
            continue
        seen.add(signature)
        candidates.append(row)
    return candidates


def create_generated_workbook(
    connection: Connection,
    *,
    preview_rows: list[dict[str, Any]],
    selected_candidate_ids: dict[int, int],
    operator_name: str,
    request_file_name: str,
) -> tuple[BytesIO, int]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Internal Quotation"

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(GENERATED_COLUMNS))
    title_cell = worksheet.cell(row=1, column=1, value="GTS Internal Quotation")
    title_cell.font = Font(bold=True, size=14, color="172033")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="EAF3F0")

    for column_index, (_, label) in enumerate(GENERATED_COLUMNS, start=1):
        cell = worksheet.cell(row=2, column=column_index, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="176D5D")

    output_row = 3
    generated_count = 0
    for preview_row in preview_rows:
        row_number = int(preview_row["row_number"])
        selected_id = selected_candidate_ids.get(row_number)
        if selected_id is None:
            continue

        request_values = preview_row["values"]
        if selected_id == BLANK_ROW_CANDIDATE_ID:
            output_values = build_blank_output_row(request_values)
        else:
            candidate = connection.execute(
                "SELECT * FROM quotation_items WHERE id = ?",
                (selected_id,),
            ).fetchone()
            if not candidate:
                continue
            product = connection.execute(
                "SELECT * FROM products WHERE id = ?",
                (candidate["product_id"],),
            ).fetchone()

            output_values = build_output_row(
                candidate,
                product,
                request_values.get("quantity"),
                request_values.get("description"),
                request_values.get("oem"),
                request_values.get("unit"),
            )
        output_values["no"] = generated_count + 1
        for column_index, (field, _) in enumerate(GENERATED_COLUMNS, start=1):
            worksheet.cell(row=output_row, column=column_index, value=output_values.get(field))
        output_row += 1
        generated_count += 1

    create_operation_log(
        connection,
        operator_name=operator_name,
        action_type="generate_quotation",
        file_name=request_file_name,
        row_count=generated_count,
        note="已生成内部报价 Excel，供立即下载。",
    )

    apply_generated_workbook_formatting(worksheet)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream, generated_count


def build_blank_output_row(request_values: dict[str, Any]) -> dict[str, Any]:
    quantity = whole_quantity(request_values.get("quantity"))
    return {
        "gts_no": request_values.get("gts_no") or None,
        "description": request_values.get("description") or None,
        "oem": request_values.get("oem") or None,
        "photo": None,
        "quantity": quantity,
        "unit": request_values.get("unit") or None,
        "total_price": None,
        "comment": request_values.get("comment") or None,
    }


def build_output_row(
    candidate: Row,
    product: Row | None,
    request_quantity: float | None,
    request_description: str | None = None,
    request_oem: str | None = None,
    request_unit: str | None = None,
) -> dict[str, Any]:
    output = {field: candidate[field] for field, _ in GENERATED_COLUMNS if field in candidate.keys()}
    if product:
        for field in ("gts_no", "description", "oem"):
            output[field] = product[field]
    output["photo"] = None
    if _has_text(output.get("updated_at")):
        output["updated_at"] = str(output["updated_at"])[:10]
    for field, request_value in (("unit", request_unit),):
        if _has_text(request_value):
            output[field] = request_value
    output["quantity"] = whole_quantity(request_quantity)
    if output["quantity"] is not None and candidate["unit_price"] is not None:
        output["total_price"] = output["quantity"] * float(candidate["unit_price"])
    elif request_quantity is None:
        output["total_price"] = None
    return output


def whole_quantity(value: Any) -> int | None:
    if value in ("", None):
        return None
    return int(round(float(value)))


def _price_key(value: Any) -> str:
    if value in ("", None):
        return ""
    return f"{float(value):.4f}"


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _has_text(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def apply_generated_workbook_formatting(worksheet) -> None:
    column_widths = {
        "A": 8,
        "B": 16,
        "C": 28,
        "D": 20,
        "E": 12,
        "F": 18,
        "G": 24,
        "H": 12,
        "I": 10,
        "J": 14,
        "K": 14,
        "L": 16,
        "M": 12,
        "N": 18,
        "O": 12,
        "P": 10,
        "Q": 10,
        "R": 10,
        "S": 20,
        "T": 18,
        "U": 18,
        "V": 28,
        "W": 16,
        "X": 22,
    }
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A3"
    currency_format = '"¥"#,##0.00'
    decimal_format = "0.00"

    field_to_column = {
        field: get_column_letter(index)
        for index, (field, _) in enumerate(GENERATED_COLUMNS, start=1)
    }
    for row in range(3, worksheet.max_row + 1):
        worksheet[f"{field_to_column['quantity']}{row}"].number_format = "0"
        worksheet[f"{field_to_column['unit_price']}{row}"].number_format = currency_format
        worksheet[f"{field_to_column['total_price']}{row}"].number_format = currency_format
        for field in ("length", "width", "height", "measurements_volume"):
            worksheet[f"{field_to_column[field]}{row}"].number_format = decimal_format
