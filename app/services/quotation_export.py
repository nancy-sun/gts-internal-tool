from __future__ import annotations

from io import BytesIO
from sqlite3 import Connection, Row
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.operation_logging import create_operation_log
from app.services.quotation_columns import BLANK_ROW_CANDIDATE_ID, GENERATED_COLUMNS
from app.services.suppliers import supplier_link_available


def create_generated_workbook(
    connection: Connection,
    *,
    preview_rows: list[dict[str, Any]],
    selected_candidate_ids: dict[int, int],
    operator_name: str,
    request_file_name: str,
) -> tuple[BytesIO, int]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Internal Quotation"
    write_generated_header(worksheet)

    output_row = 3
    generated_count = 0
    for preview_row in preview_rows:
        row_number = int(preview_row["row_number"])
        selected_id = selected_candidate_ids.get(row_number)
        if selected_id is None:
            continue

        output_values = build_selected_output_row(
            connection,
            preview_row=preview_row,
            selected_id=selected_id,
        )
        if output_values is None:
            continue
        output_values["no"] = generated_count + 1
        for column_index, (field, _) in enumerate(GENERATED_COLUMNS, start=1):
            worksheet.cell(row=output_row, column=column_index, value=output_values.get(field))
        output_row += 1
        generated_count += 1

    create_operation_log(
        connection,
        operator_name=operator_name,
        action_type="generate_quotation",
        file_name=request_file_name,
        row_count=generated_count,
        note="已生成内部报价 Excel，供立即下载。",
    )

    apply_generated_workbook_formatting(worksheet)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream, generated_count


def write_generated_header(worksheet) -> None:
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(GENERATED_COLUMNS))
    title_cell = worksheet.cell(row=1, column=1, value="GTS Internal Quotation")
    title_cell.font = Font(bold=True, size=14, color="172033")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="EAF3F0")

    for column_index, (_, label) in enumerate(GENERATED_COLUMNS, start=1):
        cell = worksheet.cell(row=2, column=column_index, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="176D5D")


def build_selected_output_row(
    connection: Connection,
    *,
    preview_row: dict[str, Any],
    selected_id: int,
) -> dict[str, Any] | None:
    request_values = preview_row["values"]
    if selected_id == BLANK_ROW_CANDIDATE_ID:
        return build_blank_output_row(request_values)

    if supplier_link_available(connection):
        candidate_query = """
            SELECT
                q.*,
                s.supplier_name
            FROM quotation_items q
            LEFT JOIN suppliers s ON s.id = q.supplier_id
            WHERE q.id = ?
        """
    else:
        candidate_query = """
            SELECT
                q.*,
                NULL AS supplier_name
            FROM quotation_items q
            WHERE q.id = ?
        """
    candidate = connection.execute(candidate_query, (selected_id,)).fetchone()
    if not candidate:
        return None
    product = connection.execute(
        "SELECT * FROM products WHERE id = ?",
        (candidate["product_id"],),
    ).fetchone()

    return build_output_row(
        candidate,
        product,
        request_values.get("quantity"),
        request_values.get("description"),
        request_values.get("oem"),
        request_values.get("unit"),
    )


def build_blank_output_row(request_values: dict[str, Any]) -> dict[str, Any]:
    quantity = whole_quantity(request_values.get("quantity"))
    return {
        "gts_no": request_values.get("gts_no") or None,
        "description": request_values.get("description") or None,
        "oem": request_values.get("oem") or None,
        "photo": None,
        "quantity": quantity,
        "unit": request_values.get("unit") or None,
        "total_price": None,
        "comment": request_values.get("comment") or None,
    }


def build_output_row(
    candidate: Row,
    product: Row | None,
    request_quantity: float | None,
    request_description: str | None = None,
    request_oem: str | None = None,
    request_unit: str | None = None,
) -> dict[str, Any]:
    output = {field: candidate[field] for field, _ in GENERATED_COLUMNS if field in candidate.keys()}
    if product:
        for field in ("gts_no", "description", "oem"):
            output[field] = product[field]
    if "supplier_name" in candidate.keys() and _has_text(candidate["supplier_name"]):
        output["factory"] = candidate["supplier_name"]
    output["photo"] = None
    if _has_text(output.get("updated_at")):
        output["updated_at"] = str(output["updated_at"])[:10]
    for field, request_value in (("unit", request_unit),):
        if _has_text(request_value):
            output[field] = request_value
    output["quantity"] = whole_quantity(request_quantity)
    if output["quantity"] is not None and candidate["unit_price"] is not None:
        output["total_price"] = output["quantity"] * float(candidate["unit_price"])
    elif request_quantity is None:
        output["total_price"] = None
    return output


def whole_quantity(value: Any) -> int | None:
    if value in ("", None):
        return None
    return int(round(float(value)))


def _has_text(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def apply_generated_workbook_formatting(worksheet) -> None:
    column_widths = {
        "A": 8,
        "B": 16,
        "C": 28,
        "D": 20,
        "E": 12,
        "F": 18,
        "G": 24,
        "H": 12,
        "I": 10,
        "J": 14,
        "K": 14,
        "L": 16,
        "M": 12,
        "N": 18,
        "O": 12,
        "P": 10,
        "Q": 10,
        "R": 10,
        "S": 20,
        "T": 18,
        "U": 18,
        "V": 28,
        "W": 16,
        "X": 22,
    }
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A3"
    currency_format = '"¥"#,##0.00'
    decimal_format = "0.00"

    field_to_column = {
        field: get_column_letter(index)
        for index, (field, _) in enumerate(GENERATED_COLUMNS, start=1)
    }
    for row in range(3, worksheet.max_row + 1):
        worksheet[f"{field_to_column['quantity']}{row}"].number_format = "0"
        worksheet[f"{field_to_column['unit_price']}{row}"].number_format = currency_format
        worksheet[f"{field_to_column['total_price']}{row}"].number_format = currency_format
        for field in ("length", "width", "height", "measurements_volume"):
            worksheet[f"{field_to_column[field]}{row}"].number_format = decimal_format
