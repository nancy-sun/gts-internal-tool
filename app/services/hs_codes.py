from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from sqlite3 import Connection, Row
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter

from app.services.excel_parser import (
    HEADER_ALIASES,
    _clean_cell_value,
    normalize_header_label,
)
from app.services.normalization import normalize_gts_no
from app.services.operation_logging import create_operation_log


HS_CODE_ALIASES = ["HS", "hscode", "海关编码", "hs code"]
HS_UPLOAD_FIELDS = ("gts_no", "hs_code")
HS_REQUEST_FIELDS = ("gts_no",)
HS_UPLOAD_LOOKUP = {
    normalize_header_label(alias): "gts_no"
    for alias in HEADER_ALIASES["gts_no"]
} | {
    normalize_header_label(alias): "hs_code"
    for alias in HS_CODE_ALIASES
}
HS_REQUEST_LOOKUP = {
    normalize_header_label(alias): "gts_no"
    for alias in HEADER_ALIASES["gts_no"]
}
MAX_ROWS = 300
HEADER_SCAN_ROWS = 100
HEADER_SCAN_COLUMNS = 80
UNMATCHED_GTS_MESSAGE = "数据库中未找到这个 GTS。"


@dataclass
class ParsedHsCodeRow:
    row_number: int
    values: dict[str, Any]
    warnings: list[str]
    errors: list[str]


def parse_hs_code_upload_workbook(workbook_path: Path) -> list[ParsedHsCodeRow]:
    return parse_hs_code_workbook(
        workbook_path,
        fields=HS_UPLOAD_FIELDS,
        header_lookup=HS_UPLOAD_LOOKUP,
        required_header_fields={"gts_no", "hs_code"},
        require_hs_code=True,
    )


def parse_hs_code_request_workbook(workbook_path: Path) -> list[ParsedHsCodeRow]:
    return parse_hs_code_workbook(
        workbook_path,
        fields=HS_REQUEST_FIELDS,
        header_lookup=HS_REQUEST_LOOKUP,
        required_header_fields={"gts_no"},
        require_hs_code=False,
    )


def parse_hs_code_workbook(
    workbook_path: Path,
    *,
    fields: tuple[str, ...],
    header_lookup: dict[str, str],
    required_header_fields: set[str],
    require_hs_code: bool,
) -> list[ParsedHsCodeRow]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        worksheet = workbook.worksheets[0]
        header_row = resolve_hs_header_row(
            worksheet,
            header_lookup,
            required_header_fields,
        )
        columns = resolve_hs_columns(worksheet, header_row, header_lookup)
        start_row = header_row + 1
        parsed_rows: list[ParsedHsCodeRow] = []
        for row_number in range(start_row, start_row + MAX_ROWS):
            values = {
                field: _clean_cell_value(worksheet[f"{columns[field]}{row_number}"].value)
                if field in columns
                else ""
                for field in fields
            }
            if all(value in ("", None) for value in values.values()):
                continue
            warnings: list[str] = []
            errors: list[str] = []
            if "hs_code" in values:
                values["hs_code"] = _text(values.get("hs_code"))
            gts_no_normalized, gts_warnings = normalize_gts_no(values["gts_no"])
            values["gts_no_normalized"] = gts_no_normalized
            if require_hs_code and (not gts_no_normalized or not values.get("hs_code")):
                continue
            warnings.extend([f"GTS {warning}" for warning in gts_warnings])
            if not gts_no_normalized:
                errors.append("GTS 不能为空。")
            parsed_rows.append(
                ParsedHsCodeRow(
                    row_number=row_number,
                    values=values,
                    warnings=warnings,
                    errors=errors,
                )
            )
        return parsed_rows
    finally:
        workbook.close()


def resolve_hs_header_row(
    worksheet,
    header_lookup: dict[str, str],
    required_fields: set[str],
) -> int:
    max_row = min(worksheet.max_row or HEADER_SCAN_ROWS, HEADER_SCAN_ROWS)
    for row in range(1, max_row + 1):
        matched_fields = set()
        for column in range(1, HEADER_SCAN_COLUMNS + 1):
            header = normalize_header_label(worksheet.cell(row=row, column=column).value)
            field = header_lookup.get(header)
            if field:
                matched_fields.add(field)
        if required_fields.issubset(matched_fields):
            return row
    return 1


def resolve_hs_columns(
    worksheet,
    header_row: int,
    header_lookup: dict[str, str],
) -> dict[str, str]:
    columns = {}
    for column_index in range(1, HEADER_SCAN_COLUMNS + 1):
        header = normalize_header_label(worksheet.cell(row=header_row, column=column_index).value)
        field = header_lookup.get(header)
        if field:
            columns.setdefault(field, get_column_letter(column_index))
    return columns


def build_hs_upload_preview(
    connection: Connection,
    parsed_rows: list[ParsedHsCodeRow],
) -> list[dict[str, Any]]:
    return build_hs_preview(
        connection,
        parsed_rows,
        unmatched_is_error=True,
        allow_historical_lookup=False,
    )


def build_hs_generate_preview(
    connection: Connection,
    parsed_rows: list[ParsedHsCodeRow],
) -> list[dict[str, Any]]:
    return build_hs_preview(
        connection,
        parsed_rows,
        unmatched_is_error=False,
        allow_historical_lookup=True,
    )


def build_hs_preview(
    connection: Connection,
    parsed_rows: list[ParsedHsCodeRow],
    *,
    unmatched_is_error: bool,
    allow_historical_lookup: bool,
) -> list[dict[str, Any]]:
    products_by_gts = fetch_products_by_gts(connection, parsed_rows)
    products_by_historical_gts = (
        fetch_products_by_historical_gts(connection, parsed_rows)
        if allow_historical_lookup
        else {}
    )
    preview_rows = []
    for parsed_row in parsed_rows:
        preview = parsed_row_to_preview(parsed_row)
        if not parsed_row.errors:
            product = products_by_gts.get(
                parsed_row.values["gts_no_normalized"],
            ) or products_by_historical_gts.get(parsed_row.values["gts_no_normalized"])
            if product:
                preview["status"] = "matched"
                preview["product"] = dict(product)
                preview["change_notices"] = build_hs_change_notices(product, parsed_row.values)
            else:
                preview["status"] = "unmatched"
                if unmatched_is_error:
                    preview["errors"].append(UNMATCHED_GTS_MESSAGE)
                else:
                    preview["warnings"].append(UNMATCHED_GTS_MESSAGE)
        preview_rows.append(preview)
    return preview_rows


def parsed_row_to_preview(parsed_row: ParsedHsCodeRow) -> dict[str, Any]:
    return {
        "row_number": parsed_row.row_number,
        "values": dict(parsed_row.values),
        "warnings": list(parsed_row.warnings),
        "errors": list(parsed_row.errors),
        "status": "invalid" if parsed_row.errors else "ready",
        "product": None,
        "change_notices": [],
    }


def fetch_products_by_gts(
    connection: Connection,
    parsed_rows: list[ParsedHsCodeRow],
) -> dict[str, Row]:
    gts_numbers = sorted(
        {
            row.values["gts_no_normalized"]
            for row in parsed_rows
            if row.values.get("gts_no_normalized")
        }
    )
    if not gts_numbers:
        return {}
    placeholders = ", ".join("?" for _ in gts_numbers)
    rows = connection.execute(
        f"""
        SELECT *
        FROM products
        WHERE gts_no_normalized IN ({placeholders})
        ORDER BY updated_at DESC, id DESC
        """,
        gts_numbers,
    ).fetchall()
    products = {}
    for row in rows:
        products.setdefault(row["gts_no_normalized"], row)
    return products


def fetch_products_by_historical_gts(
    connection: Connection,
    parsed_rows: list[ParsedHsCodeRow],
) -> dict[str, Row]:
    gts_numbers = sorted(
        {
            row.values["gts_no_normalized"]
            for row in parsed_rows
            if row.values.get("gts_no_normalized")
        }
    )
    if not gts_numbers:
        return {}
    placeholders = ", ".join("?" for _ in gts_numbers)
    rows = connection.execute(
        f"""
        SELECT
            q.gts_no_normalized AS lookup_value,
            p.*
        FROM quotation_items q
        JOIN products p ON p.id = q.product_id
        WHERE q.gts_no_normalized IN ({placeholders})
        ORDER BY q.updated_at DESC, q.id DESC
        """,
        gts_numbers,
    ).fetchall()
    products = {}
    for row in rows:
        lookup_value = _text(row["lookup_value"])
        if lookup_value and lookup_value not in products:
            products[lookup_value] = row
    return products


def build_hs_change_notices(
    product: Row | dict[str, Any],
    request_values: dict[str, Any],
) -> list[str]:
    request_gts = _text(request_values.get("gts_no"))
    if request_gts and request_values.get("gts_no_normalized") != product["gts_no_normalized"]:
        return [f"GTS 已从 {request_gts} 改为 {_text(product['gts_no']) or '(空)'}"]
    return []


def save_hs_upload_preview(
    connection: Connection,
    *,
    preview_rows: list[dict[str, Any]],
    operator_name: str,
    file_name: str,
) -> dict[str, int]:
    updated = 0
    failed = 0
    for row in preview_rows:
        if not is_valid_hs_upload_row(row):
            failed += 1
            continue
        connection.execute(
            """
            UPDATE products
            SET hs_code = ?
            WHERE id = ?
            """,
            (row["values"].get("hs_code"), row["product"]["id"]),
        )
        updated += 1

    create_operation_log(
        connection,
        operator_name=operator_name,
        action_type="update_hs_code",
        file_name=file_name,
        row_count=updated,
        note=f"更新 HS Code={updated}; 失败行={failed}",
    )
    return {"updated": updated, "failed": failed}


def is_valid_hs_upload_row(row: dict[str, Any]) -> bool:
    values = row.get("values") or {}
    product = row.get("product") or {}
    return (
        not row.get("errors")
        and row.get("status") == "matched"
        and bool(product.get("id"))
        and bool(values.get("gts_no_normalized"))
        and bool(_text(values.get("hs_code")))
    )


def create_hs_code_workbook(
    connection: Connection,
    *,
    preview_rows: list[dict[str, Any]],
    operator_name: str,
    file_name: str,
) -> tuple[BytesIO, int]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "HS Codes"
    headers = ["GTS", "OEM", "HS Code"]
    for column_index, label in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="176D5D")

    output_count = 0
    for row_index, row in enumerate(preview_rows, start=2):
        product = row.get("product") or {}
        worksheet.cell(row=row_index, column=1, value=product.get("gts_no") or row["values"].get("gts_no") or "")
        worksheet.cell(row=row_index, column=2, value=product.get("oem") or "")
        worksheet.cell(row=row_index, column=3, value=product.get("hs_code") or "")
        output_count += 1

    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 24
    worksheet.column_dimensions["C"].width = 18
    create_operation_log(
        connection,
        operator_name=operator_name,
        action_type="generate_hs_code",
        file_name=file_name,
        row_count=output_count,
        note="已生成 HS Code Excel，供立即下载。",
    )
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream, output_count


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
