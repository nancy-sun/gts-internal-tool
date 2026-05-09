import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from app.config import BASE_DIR
from app.services.normalization import normalize_gts_no, normalize_oem


QUOTATION_FIELDS = [
    "no",
    "gts_no",
    "description",
    "oem",
    "factory",
    "chinese_description",
    "quantity",
    "unit",
    "unit_price",
    "total_price",
    "item_per_package",
    "packages",
    "weight_per_package",
    "gross_weight",
    "length",
    "width",
    "height",
    "measurements_volume",
    "packaging",
    "expected_delivery",
    "comment",
]

NUMERIC_FIELDS = {"quantity", "unit_price", "total_price"}
IMPORTANT_UPLOAD_FIELDS = {"factory", "unit", "unit_price"}

HEADER_ALIASES = {
    "no": ["No.", "No", "Item No."],
    "gts_no": ["GTS No.", "GTS No", "GTS"],
    "description": ["Description", "Desc", "Desc."],
    "oem": ["OEM", "OEM.", "OEM No.", "OEM No", "OEM Number"],
    "factory": ["Factory", "工厂"],
    "chinese_description": [
        "Chinese Description",
        "Chinese Desc",
        "中文描述",
        "描述",
        "产品描述",
        "品名",
    ],
    "quantity": ["Quantity", "Qty"],
    "unit": ["Unit", "单位"],
    "unit_price": ["Unit Price", "Price", "Prix", "价格"],
    "total_price": ["Total Price", "Total", "Amount", "总价"],
    "item_per_package": [
        "Item/Package",
        "Item Per Package",
        "item/pkg",
        "Items/Package",
        "每箱数量",
    ],
    "packages": ["Packages", "Package Count", "箱数", "pkg", "pkg."],
    "weight_per_package": [
        "Weight / Package",
        "Weight Per Package",
        "Weight/Package",
        "w./pkg",
        "weight/pkg",
        "每箱重量",
    ],
    "gross_weight": ["G.W.", "G.W", "Gross Weight", "GW", "毛重"],
    "length": ["Length", "L", "L.", "长"],
    "width": ["Width", "W", "w.", "宽"],
    "height": ["Height", "H", "h.", "高"],
    "measurements_volume": [
        "Measurement",
        "Measurements",
        "Measurements / Volume",
        "Measurements (Volume)",
        "Measurement / Volume",
        "Volume",
        "mea",
        "meas",
        "meas.",
        "vol",
        "volumn",
        "vol.",
        "体积",
    ],
    "packaging": ["Packaging", "Packing", "包装"],
    "expected_delivery": ["Delivery date", "Expected Delivery", "Delivery", "Lead Time", "交期"],
    "comment": ["Comment", "Comments", "Remark", "Remarks", "备注"],
}


@dataclass
class ParsedQuotationRow:
    row_number: int
    values: dict[str, Any]
    warnings: list[str]
    errors: list[str]

    @property
    def is_valid(self) -> bool:
        return not self.errors


def load_template_config(config_path: Path | None = None) -> dict[str, Any]:
    path = config_path or BASE_DIR / "config" / "quotation_template.json"
    with path.open("r", encoding="utf-8") as config_file:
        return json.load(config_file)


def parse_full_quotation_workbook(
    workbook_path: Path,
    config: dict[str, Any] | None = None,
) -> list[ParsedQuotationRow]:
    return list(iter_full_quotation_workbook_rows(workbook_path, config))


def iter_full_quotation_workbook_rows(
    workbook_path: Path,
    config: dict[str, Any] | None = None,
):
    template = config or load_template_config()
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet_name = template.get("sheet_name")
        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        header_row = resolve_header_row(worksheet, template)
        start_row = header_row + 1
        max_rows = int(template.get("max_rows", 300))
        columns = resolve_columns(worksheet, header_row, template)
        missing_important_headers = [
            field for field in IMPORTANT_UPLOAD_FIELDS if field not in columns
        ]

        for row_number in range(start_row, start_row + max_rows):
            values = {
                field: _clean_cell_value(worksheet[f"{columns[field]}{row_number}"].value)
                if field in columns
                else ""
                for field in QUOTATION_FIELDS
            }
            if all(value in ("", None) for value in values.values()):
                continue

            warnings: list[str] = []
            errors: list[str] = []
            warnings.extend(
                f"{field_label(field)} column was not found; value will be blank."
                for field in missing_important_headers
            )
            gts_no_normalized, gts_warnings = normalize_gts_no(values["gts_no"])
            oem_normalized, oem_warnings = normalize_oem(values["oem"])
            values["gts_no_normalized"] = gts_no_normalized
            values["oem_normalized"] = oem_normalized
            warnings.extend([f"GTS No. {warning}" for warning in gts_warnings])
            warnings.extend([f"OEM {warning}" for warning in oem_warnings])

            for field in NUMERIC_FIELDS:
                values[field] = _parse_number(values[field], field, warnings)

            if not gts_no_normalized and not oem_normalized:
                errors.append("Row must contain GTS No. or OEM.")

            yield ParsedQuotationRow(
                row_number=row_number,
                values=values,
                warnings=warnings,
                errors=errors,
            )
    finally:
        workbook.close()


def resolve_header_row(worksheet, template: dict[str, Any]) -> int:
    detect_column = template.get("detect_header_from_column")
    header_label = template.get("header_label")
    if not detect_column or not header_label:
        return int(template["header_row"])

    column_index = column_index_from_string(detect_column)
    wanted = normalize_header_label(header_label)
    max_header_scan_rows = int(template.get("header_scan_rows", 100))
    max_row = worksheet.max_row or max_header_scan_rows
    scan_limit = min(max_row, max_header_scan_rows)
    for row in range(1, scan_limit + 1):
        value = worksheet.cell(row=row, column=column_index).value
        if normalize_header_label(value) == wanted:
            return row
    return int(template["header_row"])


def resolve_columns(worksheet, header_row: int, template: dict[str, Any]) -> dict[str, str]:
    detected_headers: dict[str, str] = {}
    max_scan_columns = int(template.get("header_scan_columns", 80))
    for column_index in range(1, max_scan_columns + 1):
        header_value = worksheet.cell(row=header_row, column=column_index).value
        normalized_header = normalize_header_label(header_value)
        field = HEADER_LOOKUP.get(normalized_header)
        if field:
            detected_headers.setdefault(field, get_column_letter(column_index))

    fallback_columns = template.get("columns", {})
    return {
        field: detected_headers.get(field, fallback_columns.get(field))
        for field in QUOTATION_FIELDS
        if field in detected_headers or field in fallback_columns
    }


def field_label(field: str) -> str:
    return {
        "factory": "Factory",
        "unit": "Unit",
        "unit_price": "Unit Price",
    }.get(field, field)


def normalize_header_label(value: Any) -> str:
    if value is None:
        return ""
    return "".join(character for character in str(value).upper() if character.isalnum())


HEADER_LOOKUP = {
    normalize_header_label(alias): field
    for field, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _clean_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _parse_number(value: Any, field: str, warnings: list[str]) -> float | None:
    if value in ("", None):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        warnings.append(f"{field} is not a number and will be left blank.")
        return None
