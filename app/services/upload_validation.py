from io import BytesIO
import re
from zipfile import BadZipFile

from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.services.excel_parser import (
    find_alias_header_row,
    find_configured_header_row,
    load_template_config,
    normalize_header_label,
)
from app.services.request_parser import (
    REQUEST_HEADER_LOOKUP,
    load_request_template_config,
)


def validate_xlsx_upload(upload_file: UploadFile, operator_name: str) -> str | None:
    if not operator_name.strip():
        return "请填写操作员姓名。"
    filename = upload_file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        return "只能上传 .xlsx 文件。"
    return None


def sanitize_upload_filename(
    filename: str | None,
    *,
    default: str,
    max_length: int = 120,
) -> str:
    raw_filename = (filename or default).strip() or default
    base_name = raw_filename.replace("\\", "/").split("/")[-1].strip() or default
    stem = base_name[:-5] if base_name.lower().endswith(".xlsx") else base_name
    stem = re.sub(r"[\x00-\x1f/\\:]+", "_", stem).strip(" ._") or "upload"
    max_stem_length = max(1, max_length - len(".xlsx"))
    if len(stem) > max_stem_length:
        stem = stem[:max_stem_length].rstrip(" ._") or "upload"
    return f"{stem}.xlsx"


def validate_workbook_contents(contents: bytes) -> str | None:
    return validate_workbook(contents)


def validate_full_quotation_workbook(contents: bytes) -> str | None:
    return validate_workbook(contents, validate_full_quotation_header)


def validate_request_list_workbook(contents: bytes) -> str | None:
    return validate_workbook(contents, validate_request_list_identifiers)


def validate_upload_size(
    contents: bytes,
    *,
    max_upload_size_bytes: int,
    max_upload_size_mb: int,
) -> str | None:
    if len(contents) <= max_upload_size_bytes:
        return None
    return f"上传文件超过 {max_upload_size_mb} MB。"


def validate_workbook(contents: bytes, validator=None) -> str | None:
    try:
        workbook = load_workbook(
            BytesIO(contents),
            data_only=True,
            read_only=True,
        )
    except (BadZipFile, InvalidFileException, OSError, ValueError):
        return "Excel 文件无法读取，请确认文件是有效的 .xlsx 工作簿。"

    try:
        if not workbook.worksheets:
            return "Excel 文件没有工作表，请上传包含数据的 .xlsx 文件。"
        if validator:
            return validator(workbook)
        return None
    finally:
        workbook.close()


def validate_full_quotation_header(workbook) -> str | None:
    template = load_template_config()
    sheet_name = template.get("sheet_name")
    if sheet_name and sheet_name not in workbook.sheetnames:
        return f"Excel 文件中找不到工作表：{sheet_name}。"
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    if (
        find_configured_header_row(worksheet, template) is None
        and find_alias_header_row(worksheet, template) is None
    ):
        return "系统找不到报价单表头行，请确认 Excel 中有 No.、GTS、OEM、工厂、单位、价格等表头。"
    return None


def validate_request_list_identifiers(workbook) -> str | None:
    template = load_request_template_config()
    sheet_name = template.get("sheet_name")
    if sheet_name and sheet_name not in workbook.sheetnames:
        return f"Excel 文件中找不到工作表：{sheet_name}。"
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    max_header_scan_rows = int(template.get("header_scan_rows", 100))
    max_header_scan_columns = int(template.get("header_scan_columns", 40))
    max_row = worksheet.max_row or max_header_scan_rows
    scan_row_limit = min(max_row, max_header_scan_rows)

    for row in range(1, scan_row_limit + 1):
        matched_fields = set()
        for column in range(1, max_header_scan_columns + 1):
            header = worksheet.cell(row=row, column=column).value
            field = REQUEST_HEADER_LOOKUP.get(normalize_header_label(header))
            if field:
                matched_fields.add(field)
        if matched_fields & {"gts_no", "oem"}:
            return None
    return "需求文件中找不到 GTS 或 OEM 表头，请检查需求 Excel。"
